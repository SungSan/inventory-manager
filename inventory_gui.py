"""Tkinter-based desktop interface for the inventory manager.

이 GUI는 기존 CLI 기능을 그대로 활용하면서 마우스 클릭만으로 입고/출고 및
재고 조회를 수행할 수 있도록 구성했습니다. 실행은 `python inventory_gui.py`
또는 PyInstaller로 패키징한 실행 파일을 통해 가능하며, 모든 데이터는
inventory.py와 동일한 JSON 파일을 공유합니다.
"""
from __future__ import annotations

import calendar
import json
import os
import queue
import re
import threading
import traceback
import tkinter as tk
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Dict, List, Optional, Set, Tuple

try:  # pragma: no cover - optional dependency
    from PIL import Image, ImageTk
except Exception:  # pragma: no cover - optional dependency
    Image = None
    ImageTk = None

from inventory import (
    DATA_FILE,
    Transaction,
    backup_data,
    backup_data_with_label,
    determine_artist,
    determine_category,
    ensure_period,
    export_to_xlsx,
    filter_stock_by_artist,
    format_stock_table,
    iter_history,
    load_data,
    normalize_category,
    record_transaction,
    restore_backup,
    save_data,
    summarize,
    update_stock,
)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_FILE = os.path.join(BASE_DIR, "inventory_settings.json")
FATAL_LOG = Path(BASE_DIR) / "fatal.log"
CATEGORY_LABELS = {"album": "앨범", "md": "MD"}


def apply_modern_styles(root: tk.Misc) -> None:
    """Apply a consistent, modernized ttk style for a more polished UI."""

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    base_font = ("Segoe UI", 10)
    heading_font = ("Segoe UI Semibold", 10)

    style.configure(".", font=base_font)
    style.configure("TLabel", padding=(2, 1))
    style.configure("TButton", padding=(10, 6), font=base_font)
    style.map(
        "TButton",
        background=[("active", "#2563eb")],
        foreground=[("active", "#ffffff")],
    )
    style.configure("Accent.TButton", background="#2563eb", foreground="#ffffff")
    style.map("Accent.TButton", background=[("active", "#1d4ed8")])

    style.configure("Treeview", font=base_font, rowheight=26, borderwidth=0)
    style.configure("Treeview.Heading", font=heading_font, padding=6)
    style.map(
        "Treeview",
        background=[("selected", "#e0f2fe")],
        foreground=[("selected", "#0f172a")],
    )
    style.configure("TNotebook", tabmargins=(0, 0, 0, 0))


def _log_fatal(context: str, exc: Exception) -> None:  # pragma: no cover - startup safety
    try:
        FATAL_LOG.write_text(
            f"[{datetime.now().isoformat()}] {context}: {exc!r}\n\n{traceback.format_exc()}\n",
            encoding="utf-8",
        )
    except Exception:
        pass


class AsyncSaveQueue:
    """Background saver to keep UI interactions responsive during disk writes."""

    def __init__(self) -> None:
        self._queue: "queue.Queue[tuple[Dict, bool]]" = queue.Queue()
        self._lock = threading.Lock()
        self._worker = threading.Thread(target=self._worker_loop, daemon=True)
        self._worker.start()

    def enqueue(self, data: Dict, *, update_timestamp: bool = True) -> None:
        self._queue.put((data, update_timestamp))

    def save_now(self, data: Dict, *, update_timestamp: bool = True) -> None:
        snapshot = deepcopy(data)
        with self._lock:
            save_data(snapshot, update_timestamp=update_timestamp)

    def _worker_loop(self) -> None:
        while True:
            data, update_timestamp = self._queue.get()
            # Coalesce any pending requests to write only the latest snapshot.
            try:
                while True:
                    data, update_timestamp = self._queue.get_nowait()
            except queue.Empty:
                pass
            snapshot = deepcopy(data)
            try:
                with self._lock:
                    save_data(snapshot, update_timestamp=update_timestamp)
            except Exception as exc:  # pragma: no cover - background logging only
                print(f"[AsyncSaveQueue] 저장 실패: {exc}")


def load_settings() -> Dict[str, object]:
    """Load GUI settings such as lock preferences and password."""

    defaults: Dict[str, object] = {
        "password": "",
        "lock_on_start": True,
        "lock_on_idle": True,
        "idle_minutes": 5,
        "location_presets": [],
        "google_enabled": False,
        "google_sheet_id": "",
        "google_credentials_path": "",
        "nickname": "",
        "nickname_password": "",
    }
    if not os.path.exists(SETTINGS_FILE):
        return defaults
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as fp:
            loaded = json.load(fp)
        if isinstance(loaded, dict):
            merged = defaults.copy()
            for key, value in loaded.items():
                if key == "location_presets":
                    merged[key] = value if isinstance(value, list) else []
                elif key in defaults:
                    merged[key] = value
            return merged
    except Exception:
        return defaults
    return defaults


def save_settings(settings: Dict[str, object]) -> None:
    """Persist GUI settings to disk."""

    with open(SETTINGS_FILE, "w", encoding="utf-8") as fp:
        json.dump(settings, fp, ensure_ascii=False, indent=2)


def export_stock_rows_to_xlsx(
    rows: List[Tuple[str, str, str, str, int, int, int, int, str, str]],
    per_locations: List[Tuple[str, str, str, str, str, int]],
    path: str,
) -> None:
    """Save current stock rows with opening/in/out/current metrics to Excel, plus per-location details."""

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl 모듈이 필요합니다. 'pip install openpyxl' 로 설치해 주세요.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "CurrentStock"
    ws.append(
        [
            "구분",
            "아티스트",
            "앨범/버전",
            "옵션",
            "기초재고",
            "입고합계",
            "출고합계",
            "현재고",
            "마지막 실사",
            "로케이션",
        ]
    )
    for row in rows:
        ws.append(list(row))

    detail = wb.create_sheet("Locations")
    detail.append(["구분", "아티스트", "앨범/버전", "옵션", "로케이션", "수량"])
    for category, artist, item, option, location, qty in per_locations:
        detail.append([category, artist, item, option, location, qty])
    wb.save(path)


LOCATION_MAP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "locations.json")


def load_location_entries(path: Optional[str] = None) -> List[Dict]:
    """Load map entries from locations.json (if present)."""

    path = path or LOCATION_MAP_FILE
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as fp:
            data = json.load(fp)
        if isinstance(data, list):
            return [entry for entry in data if isinstance(entry, dict)]
    except Exception:
        return []
    return []


def save_location_entries(entries: List[Dict], path: Optional[str] = None) -> None:
    """Persist map entries to locations.json with UTF-8 encoding."""

    path = path or LOCATION_MAP_FILE
    with open(path, "w", encoding="utf-8") as fp:
        json.dump(entries, fp, ensure_ascii=False, indent=2)


class CalendarPopup(tk.Toplevel):
    def __init__(self, master: tk.Misc, on_select, initial_date: Optional[date] = None):
        super().__init__(master)
        self.title("날짜 선택")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.on_select = on_select
        today = date.today()
        initial = initial_date or today
        self.current_year = initial.year
        self.current_month = initial.month
        self._build_ui()
        self._render_days()

    def _build_ui(self) -> None:
        header = ttk.Frame(self, padding=8)
        header.pack(fill=tk.X)
        ttk.Button(header, text="◀", width=3, command=self._prev_month).pack(side=tk.LEFT)
        self.month_label = ttk.Label(header, text="", width=20, anchor=tk.CENTER)
        self.month_label.pack(side=tk.LEFT, expand=True)
        ttk.Button(header, text="▶", width=3, command=self._next_month).pack(side=tk.RIGHT)

        self.days_frame = ttk.Frame(self, padding=(12, 4, 12, 12))
        self.days_frame.pack()

    def _render_days(self) -> None:
        for child in self.days_frame.winfo_children():
            child.destroy()
        month_name = f"{self.current_year}년 {self.current_month:02d}월"
        self.month_label.config(text=month_name)
        headers = ["월", "화", "수", "목", "금", "토", "일"]
        for idx, name in enumerate(headers):
            ttk.Label(self.days_frame, text=name, width=4, anchor=tk.CENTER).grid(row=0, column=idx, pady=2)
        cal = calendar.Calendar(firstweekday=0)
        for row_idx, week in enumerate(cal.monthdayscalendar(self.current_year, self.current_month), start=1):
            for col_idx, day in enumerate(week):
                if day == 0:
                    ttk.Label(self.days_frame, text="", width=4).grid(row=row_idx, column=col_idx, padx=2, pady=2)
                    continue
                btn = ttk.Button(
                    self.days_frame,
                    text=f"{day:02d}",
                    width=4,
                    command=lambda d=day: self._select_day(d),
                )
                btn.grid(row=row_idx, column=col_idx, padx=2, pady=2)

    def _select_day(self, day: int) -> None:
        chosen = date(self.current_year, self.current_month, day)
        self.on_select(chosen)
        self.destroy()

    def _prev_month(self) -> None:
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self._render_days()

    def _next_month(self) -> None:
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self._render_days()


class StockEditDialog(tk.Toplevel):
    def __init__(self, master: tk.Misc, *, item: str, location: str, quantity: int, artist: str, option: str, category: str):
        super().__init__(master)
        self.title("재고 수정")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.result: Optional[Dict[str, str]] = None

        body = ttk.Frame(self, padding=12)
        body.pack(fill=tk.BOTH, expand=True)

        ttk.Label(body, text="품목").grid(row=0, column=0, sticky=tk.W)
        self.item_var = tk.StringVar(value=item)
        ttk.Entry(body, textvariable=self.item_var, width=30).grid(row=0, column=1, sticky=tk.W)

        ttk.Label(body, text="아티스트").grid(row=1, column=0, sticky=tk.W, pady=(8, 0))
        self.artist_var = tk.StringVar(value=artist if artist != "-" else "")
        ttk.Entry(body, textvariable=self.artist_var, width=30).grid(row=1, column=1, sticky=tk.W, pady=(8, 0))

        ttk.Label(body, text="구분").grid(row=2, column=0, sticky=tk.W, pady=(8, 0))
        self.category_var = tk.StringVar(value=CATEGORY_LABELS.get(category, category))
        ttk.Combobox(body, textvariable=self.category_var, values=list(CATEGORY_LABELS.values()), state="readonly", width=12).grid(
            row=2, column=1, sticky=tk.W, pady=(8, 0)
        )

        ttk.Label(body, text="옵션").grid(row=3, column=0, sticky=tk.W, pady=(8, 0))
        self.option_var = tk.StringVar(value=option if option != "-" else "")
        ttk.Entry(body, textvariable=self.option_var, width=30).grid(row=3, column=1, sticky=tk.W, pady=(8, 0))

        ttk.Label(body, text="로케이션").grid(row=4, column=0, sticky=tk.W, pady=(8, 0))
        self.location_var = tk.StringVar(value=location)
        ttk.Entry(body, textvariable=self.location_var, width=30).grid(row=4, column=1, sticky=tk.W, pady=(8, 0))

        ttk.Label(body, text="수량").grid(row=5, column=0, sticky=tk.W, pady=(8, 0))
        self.quantity_var = tk.StringVar(value=str(quantity))
        ttk.Entry(body, textvariable=self.quantity_var, width=15).grid(row=5, column=1, sticky=tk.W, pady=(8, 0))

        btns = ttk.Frame(self, padding=(12, 0, 12, 12))
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="취소", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(btns, text="저장", command=self._save).pack(side=tk.RIGHT, padx=(0, 8))

    def _save(self) -> None:
        try:
            qty = int(self.quantity_var.get().strip())
        except ValueError:
            messagebox.showerror("오류", "수량은 정수여야 합니다.")
            return
        location = self.location_var.get().strip()
        if not location:
            messagebox.showerror("오류", "로케이션을 입력해 주세요.")
            return
        artist = self.artist_var.get().strip()
        item = self.item_var.get().strip()
        if not artist:
            messagebox.showerror("오류", "아티스트를 입력해 주세요.")
            return
        if not item:
            messagebox.showerror("오류", "품목명을 입력해 주세요.")
            return
        self.result = {
            "item": item,
            "artist": artist,
            "category": self.category_var.get(),
            "option": self.option_var.get().strip(),
            "location": location,
            "quantity": qty,
        }
        self.destroy()


class TransactionEditDialog(tk.Toplevel):
    def __init__(self, master: tk.Misc, entry: Dict[str, object], *, require_description: bool):
        super().__init__(master)
        self.title("기록 수정")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.result: Optional[Dict[str, object]] = None
        body = ttk.Frame(self, padding=12)
        body.pack(fill=tk.BOTH, expand=True)

        self.artist_var = tk.StringVar(value=entry.get("artist", ""))
        self.item_var = tk.StringVar(value=entry.get("item", ""))
        self.option_var = tk.StringVar(value=entry.get("option", ""))
        self.category_var = tk.StringVar(value=CATEGORY_LABELS.get(entry.get("category", "album"), entry.get("category", "")))
        self.location_var = tk.StringVar(value=entry.get("location", ""))
        self.quantity_var = tk.StringVar(value=str(entry.get("quantity", 1)))
        self.day_var = tk.StringVar(value=entry.get("day") or date.today().isoformat())
        self.desc_var = tk.StringVar(value=entry.get("description", ""))
        self.require_description = require_description

        labels = [
            ("아티스트", self.artist_var, 24),
            ("앨범/버전", self.item_var, 26),
            ("옵션", self.option_var, 20),
            ("구분", self.category_var, 16),
            ("로케이션", self.location_var, 20),
            ("수량", self.quantity_var, 10),
        ]
        for idx, (title, var, width) in enumerate(labels):
            ttk.Label(body, text=title).grid(row=idx, column=0, sticky=tk.W, pady=(0 if idx == 0 else 6, 0))
            ttk.Entry(body, textvariable=var, width=width).grid(row=idx, column=1, sticky=tk.W, pady=(0 if idx == 0 else 6, 0))

        ttk.Label(body, text="기록 일자").grid(row=5, column=0, sticky=tk.W, pady=(6, 0))
        date_frame = ttk.Frame(body)
        date_frame.grid(row=5, column=1, sticky=tk.W, pady=(6, 0))
        ttk.Entry(date_frame, textvariable=self.day_var, width=14, state="readonly").pack(side=tk.LEFT)
        ttk.Button(date_frame, text="달력", command=self._open_calendar).pack(side=tk.LEFT, padx=(4, 0))

        ttk.Label(body, text="상세내용").grid(row=6, column=0, sticky=tk.W, pady=(6, 0))
        ttk.Entry(body, textvariable=self.desc_var, width=40).grid(row=6, column=1, sticky=tk.W, pady=(6, 0))

        btns = ttk.Frame(self, padding=(12, 0, 12, 12))
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="취소", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(btns, text="저장", command=self._save).pack(side=tk.RIGHT, padx=(0, 8))

    def _open_calendar(self) -> None:
        CalendarPopup(self, lambda d: self.day_var.set(d.isoformat()), initial_date=date.fromisoformat(self.day_var.get()))

    def _save(self) -> None:
        artist = self.artist_var.get().strip()
        item = self.item_var.get().strip()
        option = self.option_var.get().strip()
        location = self.location_var.get().strip()
        description = self.desc_var.get().strip()
        if not all([artist, item, location]):
            messagebox.showerror("오류", "아티스트, 앨범/버전, 로케이션은 필수입니다.")
            return
        if self.require_description and not description:
            messagebox.showerror("오류", "출고 상세내용을 입력해 주세요.")
            return
        try:
            qty = int(self.quantity_var.get().strip())
            if qty <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("오류", "수량은 1 이상의 정수여야 합니다.")
            return
        day_value = self.day_var.get().strip() or date.today().isoformat()
        self.result = {
            "artist": artist,
            "item": item,
            "category": self.category_var.get(),
            "option": option,
            "location": location,
            "quantity": qty,
            "day": day_value,
            "description": description,
        }
        self.destroy()


class EventReturnDialog(tk.Toplevel):
    def __init__(self, master: tk.Misc, events: List[Tuple[int, Dict]], format_quantity) -> None:
        super().__init__(master)
        self.title("이벤트 선택")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        self.selection: Optional[int] = None
        self._vars: List[tk.BooleanVar] = []

        info = ttk.Label(self, text="어떤 이벤트 출고에 대한 입고인가요? 한 가지만 선택해 주세요.")
        info.grid(row=0, column=0, columnspan=2, padx=12, pady=(12, 6), sticky="w")

        for idx, (_, entry) in enumerate(events):
            var = tk.BooleanVar(value=False)
            self._vars.append(var)
            display = f"{entry.get('day')} | {entry.get('description','')} | {format_quantity(entry.get('quantity',0))}"
            chk = ttk.Checkbutton(
                self,
                text=display,
                variable=var,
                command=lambda i=idx: self._set_selection(i),
            )
            chk.grid(row=idx + 1, column=0, columnspan=2, padx=12, pady=2, sticky="w")

        btns = ttk.Frame(self)
        btns.grid(row=len(events) + 1, column=0, columnspan=2, pady=(10, 12))
        ttk.Button(btns, text="확인", command=self._confirm).grid(row=0, column=0, padx=6)
        ttk.Button(btns, text="취소", command=self._cancel).grid(row=0, column=1, padx=6)

    def _set_selection(self, index: int) -> None:
        for i, var in enumerate(self._vars):
            var.set(i == index)
        self.selection = index

    def _confirm(self) -> None:
        if self.selection is None:
            messagebox.showinfo("안내", "이벤트를 선택해 주세요.")
            return
        self.destroy()

    def _cancel(self) -> None:
        self.selection = None
        self.destroy()


class LocationMapEditor(tk.Toplevel):
    """Lightweight location map editor implemented with Tkinter."""

    def __init__(self, master: tk.Misc, map_path: str = LOCATION_MAP_FILE):
        super().__init__(master)
        self.title("로케이션 맵 편집기")
        self.geometry("1100x750")
        self.minsize(960, 640)
        self.transient(master)
        self.grab_set()

        self.map_store_path = map_path
        self.map_image_path: Optional[str] = None
        self.map_name_var = tk.StringVar(value="")
        self.selected: Optional[Dict] = None
        self._drag_start: Optional[Tuple[int, int]] = None
        self._mode: Optional[str] = None
        self._resize_anchor: Optional[str] = None
        self._creating_rect: Optional[int] = None
        self.entries: List[Dict] = []
        self.pan_mode = tk.BooleanVar(value=False)
        self._pan_anchor: Optional[Tuple[int, int]] = None

        outer = ttk.Frame(self, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        toolbar = ttk.Frame(outer)
        toolbar.pack(fill=tk.X)
        ttk.Button(toolbar, text="맵 이미지 불러오기", command=self._load_image).pack(side=tk.LEFT)
        ttk.Button(toolbar, text="JSON 불러오기", command=self._load_json).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(toolbar, text="JSON 저장", command=self._save_json).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Checkbutton(toolbar, text="이동 모드", variable=self.pan_mode).pack(side=tk.LEFT, padx=(12, 0))
        ttk.Label(toolbar, text="맵 이름").pack(side=tk.LEFT, padx=(12, 4))
        ttk.Entry(toolbar, textvariable=self.map_name_var, width=22).pack(side=tk.LEFT)
        ttk.Label(toolbar, textvariable=tk.StringVar(value=f"저장 위치: {self.map_store_path}"), foreground="#555").pack(
            side=tk.RIGHT
        )

        body = ttk.Frame(outer)
        body.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        canvas_wrap = ttk.Frame(body)
        canvas_wrap.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.canvas = tk.Canvas(canvas_wrap, background="#f8fafc", highlightthickness=1, highlightbackground="#d1d5db")
        self.canvas.pack(fill=tk.BOTH, expand=True)

        xscroll = ttk.Scrollbar(canvas_wrap, orient=tk.HORIZONTAL, command=self.canvas.xview)
        yscroll = ttk.Scrollbar(canvas_wrap, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        side = ttk.LabelFrame(body, text="선택 정보", padding=10)
        side.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))

        self.loc_name_var = tk.StringVar()
        self.x_var = tk.IntVar()
        self.y_var = tk.IntVar()
        self.w_var = tk.IntVar()
        self.h_var = tk.IntVar()

        ttk.Label(side, text="로케이션명").pack(anchor="w")
        ttk.Entry(side, textvariable=self.loc_name_var, width=26).pack(anchor="w", pady=(0, 6))

        for label, var in (("X", self.x_var), ("Y", self.y_var), ("Width", self.w_var), ("Height", self.h_var)):
            ttk.Label(side, text=label).pack(anchor="w")
            ttk.Entry(side, textvariable=var, width=18).pack(anchor="w", pady=(0, 4))

        btn_row = ttk.Frame(side)
        btn_row.pack(anchor="w", pady=(6, 0))
        ttk.Button(btn_row, text="값 적용", command=self._apply_fields).pack(side=tk.LEFT)
        ttk.Button(btn_row, text="선택 삭제", command=self._delete_selected).pack(side=tk.LEFT, padx=(6, 0))

        hint = ttk.Label(
            side,
            text="캔버스에서 드래그로 박스 생성\n클릭으로 선택 후 이동/리사이즈\nDelete 키로 삭제",
            foreground="#555",
        )
        hint.pack(anchor="w", pady=(10, 0))

        self.canvas.bind("<ButtonPress-1>", self._on_press)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind("<Shift-MouseWheel>", self._on_mousewheel)
        self.canvas.bind("<Delete>", lambda event: self._delete_selected())
        self.canvas.focus_set()

        self._redraw_background()

    def _load_image(self) -> None:
        path = filedialog.askopenfilename(
            parent=self,
            title="맵 이미지 선택 (PNG/JPG)",
            filetypes=[("Image", "*.png *.jpg *.jpeg"), ("PNG", "*.png"), ("JPG", "*.jpg;*.jpeg")],
        )
        if not path:
            return
        if Image is None:
            messagebox.showerror("이미지 로드 불가", "Pillow(PIL)가 필요합니다. PNG 이미지를 사용하거나 Pillow를 설치해 주세요.")
            return
        try:
            img = Image.open(path)
            self.map_image = img
            self.map_image_tk = ImageTk.PhotoImage(img)
        except Exception as exc:  # pragma: no cover - file I/O
            messagebox.showerror("이미지 로드 실패", str(exc))
            return
        self.map_image_path = path
        if not self.map_name_var.get():
            self.map_name_var.set(os.path.basename(path))
        self._redraw_background()

    def _redraw_background(self) -> None:
        self.canvas.delete("bg")
        if getattr(self, "map_image_tk", None):
            self.canvas.create_image(0, 0, image=self.map_image_tk, anchor="nw", tags="bg")
            self.canvas.configure(scrollregion=(0, 0, self.map_image_tk.width(), self.map_image_tk.height()))
        else:
            self.canvas.configure(scrollregion=(0, 0, 2000, 1200))
        for entry in self.entries:
            self._draw_entry(entry)

    def _on_press(self, event) -> None:
        self.canvas.focus_set()
        if self.pan_mode.get():
            self.canvas.scan_mark(event.x, event.y)
            self._mode = "pan"
            return

        hit = self._find_entry(event.x, event.y)
        if hit:
            self._select(hit["id"])
            x, y, w, h = hit["rect"]
            pad = 6
            self._mode = "move"
            self._resize_anchor = None
            if abs(event.x - x) <= pad:
                self._mode = "resize"
                self._resize_anchor = "w"
            elif abs(event.x - (x + w)) <= pad:
                self._mode = "resize"
                self._resize_anchor = "e"
            if abs(event.y - y) <= pad:
                self._mode = "resize"
                self._resize_anchor = (self._resize_anchor or "") + "n"
            elif abs(event.y - (y + h)) <= pad:
                self._mode = "resize"
                self._resize_anchor = (self._resize_anchor or "") + "s"
            self._drag_start = (event.x, event.y)
        else:
            self._mode = "create"
            self._drag_start = (event.x, event.y)
            self._creating_rect = self.canvas.create_rectangle(
                event.x,
                event.y,
                event.x,
                event.y,
                outline="#2563eb",
                dash=(3, 2),
                tags="temp",
            )

    def _on_drag(self, event) -> None:
        if self._mode == "pan":
            self.canvas.scan_dragto(event.x, event.y, gain=1)
        elif self._mode == "create" and self._creating_rect:
            self.canvas.coords(self._creating_rect, self._drag_start[0], self._drag_start[1], event.x, event.y)
        elif self._mode == "move" and self.selected and self._drag_start:
            dx = event.x - self._drag_start[0]
            dy = event.y - self._drag_start[1]
            self._offset_selected(dx, dy)
            self._drag_start = (event.x, event.y)
        elif self._mode == "resize" and self.selected and self._drag_start:
            self._resize_selected(event.x, event.y)

    def _on_release(self, event) -> None:
        if self._mode == "create" and self._creating_rect:
            x0, y0, x1, y1 = self.canvas.coords(self._creating_rect)
            self.canvas.delete(self._creating_rect)
            self._creating_rect = None
            self._mode = None
            if abs(x1 - x0) < 10 or abs(y1 - y0) < 10:
                return
            name = simpledialog.askstring("로케이션명", "생성할 로케이션 이름을 입력하세요.", parent=self)
            if not name:
                return
            rect = (int(min(x0, x1)), int(min(y0, y1)), int(abs(x1 - x0)), int(abs(y1 - y0)))
            entry = {"id": f"loc-{len(self.entries)+1}", "location": name, "rect": rect}
            self.entries.append(entry)
            self._draw_entry(entry)
            self._select(entry["id"])
            self._maybe_prompt_batch(entry)
        self._mode = None
        self._drag_start = None

    def _on_mousewheel(self, event) -> None:
        delta = -1 if event.delta < 0 else 1
        if event.state & 0x1:  # Shift scrolls horizontally
            self.canvas.xview_scroll(int(delta), "units")
        else:
            self.canvas.yview_scroll(int(delta), "units")

    def _find_entry(self, x: int, y: int) -> Optional[Dict]:
        for entry in reversed(self.entries):
            ex, ey, ew, eh = entry["rect"]
            if ex <= x <= ex + ew and ey <= y <= ey + eh:
                return entry
        return None

    def _select(self, entry_id: str) -> None:
        for entry in self.entries:
            entry["selected"] = entry["id"] == entry_id
            self._draw_entry(entry)
        self.selected = next((e for e in self.entries if e["id"] == entry_id), None)
        if self.selected:
            x, y, w, h = self.selected["rect"]
            self.loc_name_var.set(self.selected.get("location", ""))
            self.x_var.set(int(x))
            self.y_var.set(int(y))
            self.w_var.set(int(w))
            self.h_var.set(int(h))

    def _draw_entry(self, entry: Dict) -> None:
        if "canvas_id" in entry:
            self.canvas.delete(entry["canvas_id"])
        if "label_id" in entry:
            self.canvas.delete(entry["label_id"])
        x, y, w, h = entry["rect"]
        outline = "#2563eb" if entry.get("selected") else "#1f2937"
        entry["canvas_id"] = self.canvas.create_rectangle(x, y, x + w, y + h, outline=outline, width=2)
        entry["label_id"] = self.canvas.create_text(x + 4, y + 4, anchor="nw", text=entry.get("location", ""))

    def _offset_selected(self, dx: int, dy: int) -> None:
        if not self.selected:
            return
        x, y, w, h = self.selected["rect"]
        self.selected["rect"] = (x + dx, y + dy, w, h)
        self._draw_entry(self.selected)
        self._update_fields_from_selected()

    def _resize_selected(self, x: int, y: int) -> None:
        if not self.selected or not self._resize_anchor:
            return
        sx, sy, w, h = self.selected["rect"]
        left, top, right, bottom = sx, sy, sx + w, sy + h
        if "w" in self._resize_anchor:
            left = x
        if "e" in self._resize_anchor:
            right = x
        if "n" in self._resize_anchor:
            top = y
        if "s" in self._resize_anchor:
            bottom = y
        new_x, new_y = min(left, right), min(top, bottom)
        new_w, new_h = abs(right - left), abs(bottom - top)
        if new_w < 5 or new_h < 5:
            return
        self.selected["rect"] = (int(new_x), int(new_y), int(new_w), int(new_h))
        self._draw_entry(self.selected)
        self._update_fields_from_selected()

    def _update_fields_from_selected(self) -> None:
        if not self.selected:
            return
        x, y, w, h = self.selected["rect"]
        self.x_var.set(int(x))
        self.y_var.set(int(y))
        self.w_var.set(int(w))
        self.h_var.set(int(h))

    def _apply_fields(self) -> None:
        if not self.selected:
            return
        name = self.loc_name_var.get().strip()
        if name:
            self.selected["location"] = name
        self.selected["rect"] = (
            int(self.x_var.get()),
            int(self.y_var.get()),
            max(1, int(self.w_var.get())),
            max(1, int(self.h_var.get())),
        )
        self._draw_entry(self.selected)

    def _delete_selected(self) -> None:
        if not self.selected:
            return
        to_delete = self.selected
        self.selected = None
        self.entries = [e for e in self.entries if e["id"] != to_delete["id"]]
        for key in ("canvas_id", "label_id"):
            if key in to_delete:
                self.canvas.delete(to_delete[key])
        self.loc_name_var.set("")
        for var in (self.x_var, self.y_var, self.w_var, self.h_var):
            var.set(0)

    def _load_json(self) -> None:
        path = filedialog.askopenfilename(
            parent=self,
            title="locations.json 불러오기",
            filetypes=[("JSON", "*.json"), ("All", "*.*")],
        )
        if not path:
            return
        entries = load_location_entries(path)
        if not entries:
            messagebox.showinfo("안내", "불러올 로케이션 데이터가 없습니다.")
            return
        maps = sorted({entry.get("map", "") for entry in entries})
        chosen_map = self.map_name_var.get() or (maps[0] if maps else "")
        if len(maps) > 1:
            chosen_map = simpledialog.askstring(
                "맵 선택",
                f"불러올 맵 이름을 입력하세요. 사용 가능: {', '.join(maps)}",
                initialvalue=chosen_map,
                parent=self,
            )
            if chosen_map is None:
                return
        filtered = [e for e in entries if (e.get("map") or "") == chosen_map]
        if not filtered:
            messagebox.showinfo("안내", "선택한 맵에 대한 데이터가 없습니다.")
            return
        self.entries = []
        for idx, ent in enumerate(filtered, start=1):
            rect = ent.get("rect") or [0, 0, 0, 0]
            safe_rect = (int(rect[0]), int(rect[1]), int(rect[2]), int(rect[3])) if len(rect) == 4 else (0, 0, 0, 0)
            entry = {"id": f"loc-{idx}", "location": ent.get("location", ""), "rect": safe_rect}
            self.entries.append(entry)
        self.map_name_var.set(chosen_map)
        self._redraw_background()
        if self.entries:
            self._select(self.entries[0]["id"])

    def _save_json(self) -> None:
        if not self.entries:
            messagebox.showinfo("안내", "저장할 로케이션 박스가 없습니다.")
            return
        map_name = self.map_name_var.get().strip() or (self.map_image_path and os.path.basename(self.map_image_path))
        if not map_name:
            messagebox.showerror("맵 이름 필요", "맵 이름을 입력해 주세요.")
            return
        payload = []
        for ent in self.entries:
            x, y, w, h = ent["rect"]
            payload.append(
                {
                    "map": map_name,
                    "location": ent.get("location", ""),
                    "rect": [int(x), int(y), int(w), int(h)],
                }
            )
        save_location_entries(payload, self.map_store_path)
        messagebox.showinfo("저장 완료", f"{self.map_store_path}에 저장했습니다.")

    # ------------------------------------------------------------------ helpers
    def _maybe_prompt_batch(self, base_entry: Dict) -> None:
        if not base_entry:
            return

        def _open_dialog() -> None:
            dlg = tk.Toplevel(self)
            dlg.title("로케이션 일괄 부여")
            dlg.transient(self)
            dlg.grab_set()
            ttk.Label(dlg, text="긴 구역을 균등 분할하여 세부 로케이션을 자동 생성합니다.").grid(
                row=0, column=0, columnspan=2, padx=12, pady=(12, 4), sticky="w"
            )

            ttk.Label(dlg, text="기본 이름").grid(row=1, column=0, padx=12, pady=4, sticky="e")
            prefix_var = tk.StringVar(value=base_entry.get("location", ""))
            ttk.Entry(dlg, textvariable=prefix_var, width=26).grid(row=1, column=1, padx=12, pady=4, sticky="w")

            ttk.Label(dlg, text="세부 구간 수").grid(row=2, column=0, padx=12, pady=4, sticky="e")
            count_var = tk.IntVar(value=4)
            ttk.Spinbox(dlg, from_=2, to=50, textvariable=count_var, width=8).grid(row=2, column=1, padx=12, pady=4, sticky="w")

            ttk.Label(dlg, text="시작 번호 (자릿수 2)").grid(row=3, column=0, padx=12, pady=4, sticky="e")
            start_var = tk.IntVar(value=1)
            ttk.Spinbox(dlg, from_=1, to=999, textvariable=start_var, width=8).grid(row=3, column=1, padx=12, pady=4, sticky="w")

            ttk.Label(dlg, text="방향").grid(row=4, column=0, padx=12, pady=4, sticky="e")
            directions = {
                "좌 → 우": "LR",
                "우 → 좌": "RL",
                "상 → 하": "TB",
                "하 → 상": "BT",
            }
            dir_var = tk.StringVar(value=list(directions.keys())[0])
            dir_box = ttk.Combobox(dlg, state="readonly", values=list(directions.keys()), textvariable=dir_var)
            dir_box.grid(row=4, column=1, padx=12, pady=4, sticky="w")

            def _confirm() -> None:
                prefix = prefix_var.get().strip()
                if not prefix:
                    messagebox.showerror("이름 필요", "기본 이름을 입력해 주세요.", parent=dlg)
                    return
                try:
                    count = int(count_var.get())
                    start = int(start_var.get())
                except Exception:
                    messagebox.showerror("숫자 필요", "구간 수와 시작 번호를 숫자로 입력해 주세요.", parent=dlg)
                    return
                if count < 2:
                    messagebox.showerror("구간 수 부족", "2개 이상으로 분할해야 합니다.", parent=dlg)
                    return
                code = directions.get(dir_box.get(), "LR")
                created = self._create_segments(base_entry, prefix, count, start, code)
                if created:
                    messagebox.showinfo("완료", f"{len(created)}개의 세부 로케이션을 추가했습니다.", parent=dlg)
                dlg.destroy()

            ttk.Button(dlg, text="분할 생성", command=_confirm).grid(row=5, column=0, padx=12, pady=12, sticky="e")
            ttk.Button(dlg, text="취소", command=dlg.destroy).grid(row=5, column=1, padx=12, pady=12, sticky="w")

        if messagebox.askyesno("일괄 부여", "이 구역을 일정한 크기로 나누어 번호를 붙일까요?", parent=self):
            self.after_idle(_open_dialog)

    def _create_segments(
        self, base_entry: Dict, prefix: str, count: int, start_index: int, direction: str
    ) -> List[Dict]:
        x, y, w, h = base_entry.get("rect", (0, 0, 0, 0))
        segments: List[Dict] = []
        if direction in ("LR", "RL"):
            step = w / count
            if step < 4:
                messagebox.showerror("너비 부족", "구간이 너무 작습니다. 구간 수를 줄여 주세요.", parent=self)
                return []
            order = range(count) if direction == "LR" else reversed(range(count))
            for idx, seg in enumerate(order):
                sx = x + int(seg * step)
                sw = int(step) if seg < count - 1 else w - int(step) * (count - 1)
                name = f"{prefix}-{str(start_index + idx).zfill(2)}"
                seg_entry = {"id": f"loc-{len(self.entries)+len(segments)+1}", "location": name, "rect": (sx, y, sw, h)}
                segments.append(seg_entry)
        else:
            step = h / count
            if step < 4:
                messagebox.showerror("높이 부족", "구간이 너무 작습니다. 구간 수를 줄여 주세요.", parent=self)
                return []
            order = range(count) if direction == "TB" else reversed(range(count))
            for idx, seg in enumerate(order):
                sy = y + int(seg * step)
                sh = int(step) if seg < count - 1 else h - int(step) * (count - 1)
                name = f"{prefix}-{str(start_index + idx).zfill(2)}"
                seg_entry = {"id": f"loc-{len(self.entries)+len(segments)+1}", "location": name, "rect": (x, sy, w, sh)}
                segments.append(seg_entry)

        for ent in segments:
            self.entries.append(ent)
            self._draw_entry(ent)
        return segments


class LocationMapViewer(tk.Toplevel):
    """Display a saved map image with highlighted rectangle for a location."""

    def __init__(self, master: tk.Misc, entries: List[Dict]):
        super().__init__(master)
        self.title("로케이션 지도 보기")
        self.geometry("1180x880")
        self.minsize(960, 700)
        self.transient(master)
        self.grab_set()

        self.entries = entries
        self.index = 0
        self.scale = 1.0
        self.image = None
        self.image_tk = None
        self._overlay_rect: Optional[int] = None
        self._overlay_text: Optional[int] = None
        self._blink_job: Optional[str] = None
        self._blink_index = 0
        self._blink_palette = [
            {"fill": "#f97316", "outline": "#fb923c", "text": "#0f172a"},  # orange
            {"fill": "#22d3ee", "outline": "#0ea5e9", "text": "#111827"},  # teal
            {"fill": "#7c3aed", "outline": "#a855f7", "text": "#fdf4ff"},  # purple
        ]

        outer = ttk.Frame(self, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        top = ttk.Frame(outer)
        top.pack(fill=tk.X)
        self.caption_var = tk.StringVar(value="")
        ttk.Label(top, textvariable=self.caption_var, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        ttk.Button(top, text="이전", command=self._prev).pack(side=tk.RIGHT)
        ttk.Button(top, text="다음", command=self._next).pack(side=tk.RIGHT, padx=(6, 0))

        canvas_wrap = ttk.Frame(outer)
        canvas_wrap.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        self.canvas = tk.Canvas(canvas_wrap, background="#111827", highlightthickness=1, highlightbackground="#d1d5db")
        self.canvas.pack(fill=tk.BOTH, expand=True)
        xscroll = ttk.Scrollbar(canvas_wrap, orient=tk.HORIZONTAL, command=self.canvas.xview)
        yscroll = ttk.Scrollbar(canvas_wrap, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.canvas.bind("<MouseWheel>", self._on_wheel)
        self.canvas.bind("<Shift-MouseWheel>", self._on_wheel)
        self.canvas.bind("<ButtonPress-1>", self._start_pan)
        self.canvas.bind("<B1-Motion>", self._on_pan)
        self.canvas.bind("<ButtonPress-2>", self._start_pan)
        self.canvas.bind("<B2-Motion>", self._on_pan)
        self.canvas.bind("<ButtonRelease-1>", self._stop_pan)
        self.canvas.bind("<ButtonRelease-2>", self._stop_pan)
        self._pan_start: Optional[Tuple[int, int]] = None

        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._render()

    def _current(self) -> Dict:
        return self.entries[self.index]

    def _render(self) -> None:
        entry = self._current()
        map_path = entry.get("map") or ""
        if not os.path.isabs(map_path):
            map_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), map_path)
        if not os.path.exists(map_path):
            messagebox.showerror("맵 이미지 없음", f"{map_path} 파일을 찾을 수 없습니다.")
            self.destroy()
            return
        if Image is None:
            messagebox.showerror("Pillow 필요", "맵 이미지를 표시하려면 Pillow(PIL)가 필요합니다.")
            self.destroy()
            return
        try:
            self.image = Image.open(map_path)
        except Exception as exc:  # pragma: no cover - file I/O
            messagebox.showerror("이미지 로드 실패", str(exc))
            self.destroy()
            return
        self.scale = 1.0
        self._update_image()
        self.caption_var.set(f"{entry.get('location')} @ {os.path.basename(map_path)} ({self.index+1}/{len(self.entries)})")
        self._center_on_current()

    def _update_image(self) -> None:
        if not self.image:
            return
        resample = getattr(Image, "Resampling", None)
        method = resample.LANCZOS if resample else Image.LANCZOS
        scaled = self.image.resize(
            (int(self.image.width * self.scale), int(self.image.height * self.scale)),
            resample=method,
        )
        self.image_tk = ImageTk.PhotoImage(scaled)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, image=self.image_tk, anchor="nw", tags="bg")
        self.canvas.configure(scrollregion=(0, 0, scaled.width, scaled.height))
        self._draw_overlay()

    def _draw_overlay(self) -> None:
        entry = self._current()
        rect = entry.get("rect") or [0, 0, 0, 0]
        if len(rect) != 4:
            return
        self._cancel_blink()
        x, y, w, h = rect
        sx = x * self.scale
        sy = y * self.scale
        sw = w * self.scale
        sh = h * self.scale
        self.canvas.delete("overlay")
        self._overlay_rect = self.canvas.create_rectangle(
            sx,
            sy,
            sx + sw,
            sy + sh,
            outline="#f97316",
            width=4,
            fill="#f97316",
            stipple="gray25",
            tags="overlay",
        )
        self._overlay_text = self.canvas.create_text(
            sx + sw / 2, sy + sh / 2, text="여기입니다", fill="#0f172a", font=("Arial", 13, "bold"), tags="overlay"
        )
        self._blink_index = 0
        self._start_blink()

    def _start_blink(self) -> None:
        self._cancel_blink()
        # faster cycle for improved visibility
        self._blink_job = self.after(300, self._blink)

    def _blink(self) -> None:
        if not self._overlay_rect or not self.winfo_exists():
            return
        self._blink_index = (self._blink_index + 1) % len(self._blink_palette)
        palette = self._blink_palette[self._blink_index]
        self.canvas.itemconfigure(
            self._overlay_rect,
            fill=palette["fill"],
            outline=palette["outline"],
            stipple="gray25",
        )
        if self._overlay_text:
            self.canvas.itemconfigure(self._overlay_text, fill=palette["text"])
        self._blink_job = self.after(300, self._blink)

    def _cancel_blink(self) -> None:
        if self._blink_job is not None:
            try:
                self.after_cancel(self._blink_job)
            except Exception:
                pass
            self._blink_job = None

    def _center_on_current(self) -> None:
        entry = self._current()
        rect = entry.get("rect") or [0, 0, 0, 0]
        if len(rect) != 4 or not self.image_tk:
            return

        def _do_center() -> None:
            canvas_w = max(1, self.canvas.winfo_width())
            canvas_h = max(1, self.canvas.winfo_height())
            x, y, w, h = rect
            sx = x * self.scale
            sy = y * self.scale
            sw = w * self.scale
            sh = h * self.scale
            cx = sx + sw / 2
            cy = sy + sh / 2
            total_w = max(self.image_tk.width(), canvas_w)
            total_h = max(self.image_tk.height(), canvas_h)

            def clamp(val: float) -> float:
                return max(0.0, min(1.0, val))

            self.canvas.xview_moveto(clamp((cx - canvas_w / 2) / total_w))
            self.canvas.yview_moveto(clamp((cy - canvas_h / 2) / total_h))

        # Canvas sizes are valid only after it is drawn.
        self.after_idle(_do_center)

    def _on_wheel(self, event) -> None:
        if event.state & 0x4:  # Ctrl + 휠은 확대/축소
            self._on_zoom(event)
            return
        step = -1 if event.delta < 0 else 1
        if event.state & 0x1:
            self.canvas.xview_scroll(int(step), "units")
        else:
            self.canvas.yview_scroll(int(step), "units")

    def _on_zoom(self, event) -> None:
        if Image is None:
            return
        delta = 1.1 if event.delta > 0 else 0.9
        self.scale = max(0.2, min(5.0, self.scale * delta))
        self._update_image()

    def _start_pan(self, event) -> None:
        self._pan_start = (event.x, event.y)

    def _on_pan(self, event) -> None:
        if not self._pan_start:
            return
        dx = self._pan_start[0] - event.x
        dy = self._pan_start[1] - event.y
        self.canvas.xview_scroll(int(dx / 2), "units")
        self.canvas.yview_scroll(int(dy / 2), "units")
        self._pan_start = (event.x, event.y)

    def _stop_pan(self, _event=None) -> None:
        self._pan_start = None

    def _next(self) -> None:
        self.index = (self.index + 1) % len(self.entries)
        self._render()

    def _prev(self) -> None:
        self.index = (self.index - 1) % len(self.entries)
        self._render()

    def _on_close(self) -> None:
        self._cancel_blink()
        self.destroy()


class InventoryApp:
    """Simple desktop window that wraps the CLI helpers with Tkinter widgets."""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Inventory Manager")
        self.settings = load_settings()
        self.data = load_data()
        self.data.setdefault("activity_log", [])
        if self.data.get("last_load_error"):
            err = self.data["last_load_error"]
            messagebox.showwarning(
                "데이터 복구",
                "기존 데이터 파일을 불러올 수 없어 새 데이터로 시작합니다.\n"
                f"오류: {err.get('error')}\n"
                f"백업: {err.get('corrupt_backup')}",
            )
        self._save_queue = AsyncSaveQueue()
        self._activity_cache: Dict[Tuple[str, str, int, Optional[str]], Dict[Tuple[str, str], Dict[str, int]]] = {}
        self.history_cache: Dict[str, List[Dict]] = {"in": [], "out": []}
        self.history_indices: Dict[str, List[int]] = {"in": [], "out": []}
        self.stock_rows: List[Tuple[str, str, str, int, int, int, int, str, str]] = []
        self.stock_row_lookup: Dict[str, Dict[str, object]] = {}
        self.checked_stock_ids: set[str] = set()
        self.audit_counts: Dict[str, int] = {}
        self.audit_entry_map: Dict[str, Dict[str, object]] = {}
        self.audit_window: Optional[tk.Toplevel] = None
        self.audit_count_var = tk.StringVar()
        self.tx_date_var = tk.StringVar(value=date.today().isoformat())
        self.tx_detail_var = tk.StringVar()
        self.option_var = tk.StringVar()
        self.event_var = tk.BooleanVar(value=False)
        self.history_start_var = tk.StringVar()
        self.history_end_var = tk.StringVar()
        self.history_artist_var = tk.StringVar(value="전체")
        self.current_history_type = "in"
        self.history_event_filter = False
        self._hover_location_row: Optional[str] = None
        self._event_sessions: Dict[str, str] = {}
        self._lock_window: Optional[tk.Toplevel] = None
        self.last_activity = datetime.now()
        self._idle_job: Optional[str] = None
        self._build_layout()
        self._apply_location_presets()
        self._bind_activity_hooks()
        self._refresh_artist_options()
        self.refresh_stock()
        self._initialize_history_defaults()
        self._start_idle_watch()
        self._maybe_lock_on_start()

    # ------------------------------------------------------------------ UI setup
    def _build_layout(self) -> None:
        self.root.geometry("1920x1080")
        main = ttk.Frame(self.root, padding=16)
        main.pack(fill=tk.BOTH, expand=True)

        self.status_var = tk.StringVar(value="재고 데이터를 불러왔습니다.")

        toolbar = ttk.Frame(main)
        toolbar.pack(fill=tk.X, pady=(0, 8))
        ttk.Button(toolbar, text="⚙ 설정", command=self.open_settings_window).pack(side=tk.LEFT)
        ttk.Button(toolbar, text="Inventory Manager로 업데이트", command=self.sync_google_drive).pack(
            side=tk.LEFT, padx=(8, 0)
        )
        ttk.Button(toolbar, text="구글 드라이브로 업데이트", command=self.upload_google_drive).pack(
            side=tk.LEFT, padx=(8, 0)
        )
        ttk.Label(toolbar, text="잠금이 필요한 경우 설정에서 비밀번호와 시간을 지정하세요.", foreground="#555").pack(
            side=tk.LEFT, padx=(8, 0)
        )

        self._build_entry_section(main)

        content = ttk.Frame(main)
        content.pack(fill=tk.BOTH, expand=True)

        left = ttk.Frame(content)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right = ttk.Frame(content)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))

        self._build_stock_section(left)
        self._build_history_section(right)

        self.status_label = ttk.Label(main, textvariable=self.status_var, foreground="#1c6b1c")
        self.status_label.pack(anchor="w", pady=(8, 0))

    # ------------------------------------------------------------------ persistence helpers
    def _mark_last_updated(self) -> None:
        self.data["last_updated"] = datetime.now().isoformat()

    def _save_async(self) -> None:
        self._mark_last_updated()
        self._save_queue.enqueue(self.data, update_timestamp=False)

    def _save_now(self) -> None:
        self._mark_last_updated()
        self._save_queue.save_now(self.data, update_timestamp=False)

    def _bind_activity_hooks(self) -> None:
        """Track user input to keep the idle timer accurate."""

        for sequence in ("<Motion>", "<KeyPress>", "<ButtonPress>"):
            self.root.bind_all(sequence, self._update_activity, add=True)

    def _normalize_category(self, value: str) -> str:
        return normalize_category(value)

    def _category_label(self, value: str) -> str:
        key = self._normalize_category(value)
        return CATEGORY_LABELS.get(key, key or "앨범")

    def _open_location_picker(self) -> None:
        options: Set[str] = set(self.settings.get("location_presets", []))
        stock = self.data.get("stock", {})
        for item_locations in stock.values():
            for option_locations in item_locations.values():
                options.update(option_locations.keys())
        sorted_options = sorted(opt for opt in options if opt)
        win = tk.Toplevel(self.root)
        win.title("로케이션 선택")
        win.geometry("320x360")
        win.transient(self.root)
        win.grab_set()

        filter_var = tk.StringVar()

        def refresh_list() -> None:
            needle = filter_var.get().strip().lower()
            listbox.delete(0, tk.END)
            for opt in sorted_options:
                if not needle or needle in opt.lower():
                    listbox.insert(tk.END, opt)

        def choose_from_selection(event=None) -> None:
            if not listbox.curselection():
                return
            choice = listbox.get(listbox.curselection()[0])
            self.location_var.set(choice)
            win.destroy()

        top = ttk.Frame(win, padding=8)
        top.pack(fill=tk.X)
        ttk.Label(top, text="필터").pack(side=tk.LEFT)
        entry = ttk.Entry(top, textvariable=filter_var)
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))
        entry.bind("<KeyRelease>", lambda _e: refresh_list())

        listbox = tk.Listbox(win, activestyle="dotbox")
        listbox.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        listbox.bind("<Double-Button-1>", choose_from_selection)

        btns = ttk.Frame(win, padding=8)
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="선택", command=choose_from_selection).pack(side=tk.RIGHT)
        ttk.Button(btns, text="닫기", command=win.destroy).pack(side=tk.RIGHT, padx=(0, 8))

        refresh_list()
        entry.focus_set()

    def _build_entry_section(self, parent: ttk.Frame) -> None:
        box = ttk.LabelFrame(parent, text="입/출고 기록", padding=12)
        box.pack(fill=tk.X, pady=(0, 12))

        self.artist_var = tk.StringVar()
        self.category_var = tk.StringVar(value="앨범")
        self.item_var = tk.StringVar()
        self.location_var = tk.StringVar()
        self.quantity_var = tk.StringVar(value="1")

        row = ttk.Frame(box)
        row.pack(fill=tk.X, pady=4)
        ttk.Label(row, text="아티스트").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(row, textvariable=self.artist_var, width=20).grid(row=1, column=0, padx=(0, 10))
        ttk.Label(row, text="구분").grid(row=0, column=1, sticky=tk.W)
        ttk.Combobox(row, textvariable=self.category_var, values=["앨범", "MD"], width=8, state="readonly").grid(
            row=1, column=1, padx=(0, 10)
        )
        ttk.Label(row, text="앨범/버전").grid(row=0, column=2, sticky=tk.W)
        ttk.Entry(row, textvariable=self.item_var, width=22).grid(row=1, column=2, padx=(0, 10))
        ttk.Label(row, text="옵션").grid(row=0, column=3, sticky=tk.W)
        ttk.Entry(row, textvariable=self.option_var, width=18).grid(row=1, column=3, padx=(0, 10))
        ttk.Label(row, text="로케이션").grid(row=0, column=4, sticky=tk.W)
        loc_frame = ttk.Frame(row)
        loc_frame.grid(row=1, column=4, padx=(0, 10), sticky=tk.W)
        self.location_combo = ttk.Combobox(
            loc_frame,
            textvariable=self.location_var,
            width=14,
            values=self.settings.get("location_presets", []),
            state="normal",
        )
        self.location_combo.pack(side=tk.LEFT)
        ttk.Button(loc_frame, text="목록", command=self._open_location_picker).pack(side=tk.LEFT, padx=(4, 0))

        ttk.Label(row, text="수량").grid(row=0, column=5, sticky=tk.W)
        ttk.Entry(row, textvariable=self.quantity_var, width=10).grid(row=1, column=5, padx=(0, 10))
        ttk.Label(row, text="기록 일자").grid(row=0, column=6, sticky=tk.W)
        date_frame = ttk.Frame(row)
        date_frame.grid(row=1, column=6)
        ttk.Entry(date_frame, textvariable=self.tx_date_var, width=14, state="readonly").pack(side=tk.LEFT)
        ttk.Button(date_frame, text="달력", command=self.open_tx_calendar).pack(side=tk.LEFT, padx=(4, 0))

        desc_row = ttk.Frame(box)
        desc_row.pack(fill=tk.X, pady=(6, 2))
        ttk.Label(desc_row, text="상세 내용").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(desc_row, textvariable=self.tx_detail_var, width=80).grid(row=1, column=0, sticky=tk.W)
        ttk.Checkbutton(desc_row, text="이벤트 출고", variable=self.event_var).grid(row=1, column=1, padx=(10, 0))

        helper = ttk.Label(
            box,
            text="기록 일자를 비우면 오늘 날짜로 저장되며, 수량/로케이션/아티스트/앨범을 모두 입력해 주세요.",
            foreground="#555",
        )
        helper.pack(anchor="w", pady=(4, 0))

        btn_row = ttk.Frame(box)
        btn_row.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(btn_row, text="입고 기록", command=lambda: self.submit_transaction("in"), width=20).pack(
            side=tk.LEFT
        )
        ttk.Button(btn_row, text="출고 기록", command=lambda: self.submit_transaction("out"), width=20).pack(
            side=tk.LEFT, padx=8
        )
        ttk.Button(btn_row, text="데이터 새로고침", command=self.reload_data).pack(side=tk.LEFT)
        ttk.Button(btn_row, text="로케이션 맵 편집기", command=self._open_location_map_editor).pack(side=tk.RIGHT)

    def _build_stock_section(self, parent: ttk.Frame) -> None:
        box = ttk.LabelFrame(parent, text="현재 재고", padding=12)
        box.pack(fill=tk.BOTH, expand=True)

        filter_row = ttk.Frame(box)
        filter_row.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(filter_row, text="아티스트 필터").pack(side=tk.LEFT)
        self.artist_filter_var = tk.StringVar(value="")
        self.artist_filter = ttk.Combobox(filter_row, textvariable=self.artist_filter_var, width=25, state="readonly")
        self.artist_filter.bind("<<ComboboxSelected>>", lambda event: self.refresh_stock())
        self.artist_filter.pack(side=tk.LEFT, padx=8)
        ttk.Button(filter_row, text="초기화", command=self.clear_artist_filter).pack(side=tk.LEFT)

        ttk.Label(filter_row, text="구분").pack(side=tk.LEFT, padx=(16, 0))
        self.category_filter_var = tk.StringVar(value="전체")
        self.category_filter = ttk.Combobox(
            filter_row, textvariable=self.category_filter_var, width=12, state="readonly"
        )
        self.category_filter["values"] = ["전체", "앨범", "MD"]
        self.category_filter.current(0)
        self.category_filter.bind("<<ComboboxSelected>>", lambda _e: self.refresh_stock())
        self.category_filter.pack(side=tk.LEFT, padx=8)
        ttk.Button(filter_row, text="초기화", command=lambda: self.category_filter_var.set("전체") or self.refresh_stock()).pack(
            side=tk.LEFT
        )

        ttk.Label(filter_row, text="로케이션 필터").pack(side=tk.LEFT, padx=(16, 0))
        self.location_filter_var = tk.StringVar(value="전체")
        self.location_filter = ttk.Combobox(
            filter_row, textvariable=self.location_filter_var, width=20, state="readonly"
        )
        self.location_filter.bind("<<ComboboxSelected>>", lambda event: self.refresh_stock())
        self.location_filter["values"] = ["전체"]
        self.location_filter.current(0)
        self.location_filter.pack(side=tk.LEFT, padx=8)
        ttk.Button(filter_row, text="초기화", command=self.clear_location_filter).pack(side=tk.LEFT)

        columns = (
            "select",
            "audit_age",
            "category",
            "artist",
            "item",
            "option",
            "opening",
            "in_total",
            "out_total",
            "qty",
            "location",
        )
        self.stock_tree = ttk.Treeview(box, columns=columns, show="headings", height=18, selectmode="browse")
        headings = [
            ("select", "선택", 60, tk.CENTER),
            ("audit_age", "실사 경과", 0, tk.CENTER),
            ("category", "구분", 90, tk.CENTER),
            ("artist", "아티스트", 140, tk.CENTER),
            ("item", "앨범/버전", 200, tk.CENTER),
            ("option", "옵션", 120, tk.CENTER),
            ("opening", "기초재고", 0, tk.CENTER),
            ("in_total", "입고합계", 0, tk.CENTER),
            ("out_total", "출고합계", 0, tk.CENTER),
            ("qty", "현재고", 80, tk.CENTER),
            ("location", "로케이션", 150, tk.CENTER),
        ]
        for col, title, width, anchor in headings:
            self.stock_tree.heading(col, text=title)
            if width == 0:
                self.stock_tree.column(col, width=0, minwidth=0, stretch=False, anchor=anchor)
            else:
                self.stock_tree.column(col, width=width, anchor=anchor)
        self.stock_tree["displaycolumns"] = ("select", "category", "artist", "item", "option", "qty", "location")
        self.stock_tree.tag_configure("negative", foreground="#b91c1c")
        self.stock_tree.tag_configure("audit_recent", foreground="#166534", background="#ecfdf3")
        self.stock_tree.tag_configure("audit_mid", foreground="#92400e", background="#fffbeb")
        self.stock_tree.tag_configure("audit_old", foreground="#991b1b", background="#fef2f2")
        self.stock_tree.tag_configure("audit_partial", background="#fff7ed", foreground="#92400e")
        self.stock_tree.tag_configure("location_hover", background="#e5edff")
        self.stock_tree.tag_configure("even", background="#f8fafc")
        self.stock_tree.tag_configure("odd", background="#ffffff")
        self.stock_tree.pack(fill=tk.BOTH, expand=True)
        self.stock_tree.bind("<Double-1>", lambda event: self.fill_transaction_from_stock())
        self.stock_tree.bind("<Button-1>", self._on_stock_click)
        self.stock_tree.bind("<Motion>", self._on_stock_motion)
        self.stock_tree.bind("<Leave>", lambda _: self._clear_location_hover())

        scrollbar = ttk.Scrollbar(box, orient=tk.VERTICAL, command=self.stock_tree.yview)
        self.stock_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        xscroll = ttk.Scrollbar(box, orient=tk.HORIZONTAL, command=self.stock_tree.xview)
        self.stock_tree.configure(xscrollcommand=xscroll.set)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)

        action_row = ttk.Frame(box)
        action_row.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(action_row, text="선택 재고 수정", command=self.edit_selected_stock).pack(side=tk.LEFT)
        ttk.Button(action_row, text="현재 재고 엑셀 저장", command=self.export_stock).pack(side=tk.LEFT, padx=8)
        ttk.Button(action_row, text="선택 재고 삭제", command=self.delete_selected_stock).pack(side=tk.LEFT)
        ttk.Button(action_row, text="재고 실사 모드", command=self.open_stock_audit).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(action_row, text="백업 불러오기", command=self.restore_from_backup_file).pack(side=tk.RIGHT)
        ttk.Button(action_row, text="데이터 백업", command=self.backup_now).pack(side=tk.RIGHT, padx=(0, 8))

    def _build_history_section(self, parent: ttk.Frame) -> None:
        box = ttk.LabelFrame(parent, text="입출고 검색", padding=12)
        box.pack(fill=tk.BOTH, expand=True)

        form = ttk.Frame(box)
        form.pack(fill=tk.X)
        self.history_artist_var = tk.StringVar()
        ttk.Label(form, text="시작 일자").grid(row=0, column=0, sticky=tk.W)
        start_frame = ttk.Frame(form)
        start_frame.grid(row=1, column=0, padx=(0, 8), sticky=tk.W)
        ttk.Entry(start_frame, textvariable=self.history_start_var, width=12, state="readonly").pack(side=tk.LEFT)
        ttk.Button(start_frame, text="달력", command=lambda: self.open_history_calendar(self.history_start_var)).pack(
            side=tk.LEFT, padx=(4, 0)
        )
        ttk.Label(form, text="종료 일자").grid(row=0, column=1, sticky=tk.W)
        end_frame = ttk.Frame(form)
        end_frame.grid(row=1, column=1, padx=(0, 8), sticky=tk.W)
        ttk.Entry(end_frame, textvariable=self.history_end_var, width=12, state="readonly").pack(side=tk.LEFT)
        ttk.Button(end_frame, text="달력", command=lambda: self.open_history_calendar(self.history_end_var)).pack(
            side=tk.LEFT, padx=(4, 0)
        )
        ttk.Label(form, text="아티스트 필터").grid(row=0, column=2, sticky=tk.W)
        self.history_artist_combo = ttk.Combobox(
            form, textvariable=self.history_artist_var, width=20, state="readonly"
        )
        self.history_artist_combo.grid(row=1, column=2, padx=(0, 8))
        self.history_artist_combo.bind(
            "<<ComboboxSelected>>",
            lambda _event: self.search_history(self.current_history_type or "in", triggered_by_calendar=True),
        )

        btns = ttk.Frame(box)
        btns.pack(fill=tk.X, pady=8)
        for i in range(6):
            btns.columnconfigure(i, weight=1)
        ttk.Button(btns, text="입고 검색", command=lambda: self.search_history("in")).grid(row=0, column=0, padx=4, pady=2, sticky="w")
        ttk.Button(btns, text="출고 검색", command=lambda: self.search_history("out")).grid(row=0, column=1, padx=4, pady=2, sticky="w")
        ttk.Button(
            btns,
            text="이벤트 출고(미해결)",
            command=lambda: self.search_history("out", event_only=True, event_open_only=True),
        ).grid(row=0, column=2, padx=4, pady=2, sticky="w")
        ttk.Button(btns, text="입고 엑셀 저장", command=lambda: self.export_history("in")).grid(row=0, column=3, padx=4, pady=2, sticky="w")
        ttk.Button(btns, text="출고 엑셀 저장", command=lambda: self.export_history("out")).grid(row=0, column=4, padx=4, pady=2, sticky="w")
        ttk.Button(btns, text="이벤트 색인 해제", command=self.clear_event_flag).grid(row=1, column=0, padx=4, pady=2, sticky="w")
        ttk.Button(btns, text="선택 내역 삭제", command=self.delete_history_entry).grid(row=1, column=5, padx=4, pady=2, sticky="e")

        self.history_caption = tk.StringVar(value="최근 입고 결과")
        ttk.Label(box, textvariable=self.history_caption, anchor="w").pack(fill=tk.X, pady=(0, 4))

        columns = ("day", "artist", "category", "item", "option", "location", "quantity", "description")
        self.history_tree = ttk.Treeview(box, columns=columns, show="headings", height=10)
        self.history_tree.heading("day", text="일자")
        self.history_tree.heading("artist", text="아티스트")
        self.history_tree.heading("category", text="구분")
        self.history_tree.heading("item", text="앨범/버전")
        self.history_tree.heading("option", text="옵션")
        self.history_tree.heading("location", text="로케이션")
        self.history_tree.heading("quantity", text="수량")
        self.history_tree.heading("description", text="상세내용")
        self.history_tree.column("day", width=90, anchor=tk.CENTER)
        self.history_tree.column("artist", width=140, anchor=tk.CENTER)
        self.history_tree.column("category", width=90, anchor=tk.CENTER)
        self.history_tree.column("item", width=180, anchor=tk.CENTER)
        self.history_tree.column("option", width=120, anchor=tk.CENTER)
        self.history_tree.column("location", width=110, anchor=tk.CENTER)
        self.history_tree.column("quantity", width=60, anchor=tk.CENTER)
        self.history_tree.column("description", width=220, anchor=tk.CENTER)
        self.history_tree.tag_configure("event_out", background="#fff3cd")
        self.history_tree.tag_configure("even", background="#f8fafc")
        self.history_tree.tag_configure("odd", background="#ffffff")
        self.history_tree.pack(fill=tk.BOTH, expand=True)
        self.history_tree.bind("<Double-1>", self.edit_history_entry)

        scroll = ttk.Scrollbar(box, orient=tk.VERTICAL, command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        xscroll = ttk.Scrollbar(box, orient=tk.HORIZONTAL, command=self.history_tree.xview)
        self.history_tree.configure(xscrollcommand=xscroll.set)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)

    # ------------------------------------------------------------------ helpers
    def set_status(self, message: str, *, error: bool = False) -> None:
        color = "#b91c1c" if error else "#1c6b1c"
        self.status_var.set(message)
        self.status_label.configure(foreground=color)

    def _current_actor(self) -> str:
        nickname = str(self.settings.get("nickname", "")).strip()
        return nickname or "미지정 사용자"

    def _log_user_action(self, action: str, *, persist: bool = False) -> None:
        entry = {
            "actor": self._current_actor(),
            "action": action,
            "timestamp": datetime.now().isoformat(timespec="seconds"),
        }
        log = self.data.setdefault("activity_log", [])
        log.append(entry)
        if len(log) > 5000:
            self.data["activity_log"] = log[-5000:]
        if persist:
            self._save_async()

    @staticmethod
    def _format_location_detail(locations: Dict[str, int]) -> str:
        if not locations:
            return "-"
        parts = [f"{loc}({qty:,})" for loc, qty in sorted(locations.items())]
        if len(parts) == 1:
            return parts[0]
        return f"{len(parts)}곳: " + ", ".join(parts)

    @staticmethod
    def _format_quantity(value: int) -> str:
        return f"{value:,}"

    def _initialize_history_defaults(self) -> None:
        start, end = self._default_history_range()
        self.history_start_var.set(start)
        self.history_end_var.set(end)
        self.history_artist_var.set("전체")
        self.history_event_filter = False
        self.search_history("in", triggered_by_calendar=True)
        self.search_history("out", triggered_by_calendar=True)
        self.search_history("in", triggered_by_calendar=True)

    def reload_data(self) -> None:
        self.data = load_data()
        self._activity_cache = {}
        self.history_cache = {"in": [], "out": []}
        self.history_indices = {"in": [], "out": []}
        self.current_history_type = "in"
        self._refresh_artist_options()
        self.refresh_stock()
        self._initialize_history_defaults()
        self.set_status("데이터를 다시 불러왔습니다.")

    def backup_now(self) -> None:
        try:
            backup_data(self.data)
        except Exception as exc:  # pragma: no cover - UI feedback path
            messagebox.showerror("오류", f"백업에 실패했습니다: {exc}")
            return
        messagebox.showinfo("완료", "현재 데이터를 백업 폴더에 저장했습니다.")

    def restore_from_backup_file(self) -> None:
        path = filedialog.askopenfilename(
            title="백업 파일 선택",
            filetypes=[("JSON Files", "*.json"), ("All Files", "*")],
            initialdir=DATA_FILE.parent / "backups",
        )
        if not path:
            return
        try:
            self.data = restore_backup(path)
        except Exception as exc:  # pragma: no cover - UI feedback path
            messagebox.showerror("오류", f"백업 불러오기 실패: {exc}")
            return
        self._refresh_artist_options()
        self.refresh_stock()
        self.search_history("in", triggered_by_calendar=True)
        self.search_history("out", triggered_by_calendar=True)
        self.search_history(self.current_history_type or "in", triggered_by_calendar=True)
        self.set_status("백업을 불러왔습니다.")

    def _open_location_map_editor(self) -> None:
        if Image is None:
            messagebox.showinfo(
                "안내",
                "맵 이미지를 다루려면 Pillow(PIL) 라이브러리가 필요합니다. 설치 후 다시 시도해 주세요.",
            )
            return
        LocationMapEditor(self.root, LOCATION_MAP_FILE)

    def _refresh_artist_options(self) -> None:
        metadata = self.data.get("item_metadata", {})
        artists = sorted({info.get("artist") for info in metadata.values() if info.get("artist")})
        self.artist_filter["values"] = ["전체"] + artists if artists else ["전체"]
        self.artist_filter.current(0)
        if hasattr(self, "history_artist_combo"):
            options = ["전체"] + artists if artists else ["전체"]
            self.history_artist_combo["values"] = options
            if self.history_artist_var.get() not in options:
                self.history_artist_var.set("전체")
            if not self.history_artist_combo.get():
                self.history_artist_combo.current(0)

    def _refresh_location_filter_options(self, locations: List[str]) -> None:
        if not hasattr(self, "location_filter"):
            return
        options = ["전체"] + [loc for loc in locations if loc]
        current = self.location_filter_var.get()
        self.location_filter["values"] = options if options else ["전체"]
        if current not in self.location_filter["values"]:
            self.location_filter.current(0)

    def _apply_location_presets(self) -> None:
        presets = [loc for loc in self.settings.get("location_presets", []) if isinstance(loc, str)]
        if hasattr(self, "location_combo"):
            self.location_combo["values"] = presets
        if hasattr(self, "location_filter"):
            # merge presets into filter options while keeping current data-driven locations
            existing = list(self.location_filter["values"]) if self.location_filter["values"] else ["전체"]
            merged = ["전체"] + sorted({v for v in existing if v != "전체"} | set(presets))
            current = self.location_filter_var.get()
            self.location_filter["values"] = merged
            if current not in merged:
                self.location_filter.current(0)

    def clear_artist_filter(self) -> None:
        self.artist_filter.current(0)
        self.refresh_stock()

    def clear_location_filter(self) -> None:
        if hasattr(self, "location_filter"):
            if self.location_filter["values"]:
                self.location_filter.current(0)
        self.refresh_stock()

    def _audit_status_for_row(self, item: str, option: str, locations: Dict[str, int]) -> Tuple[str, str, bool]:
        """Return (label, tag, partial) for the last audit time of the given item/option across locations."""

        metadata = self.data.get("item_metadata", {})
        info = metadata.get(item, {}) if isinstance(metadata.get(item, {}), dict) else {}
        audits = info.get("last_audit", {}) if isinstance(info.get("last_audit", {}), dict) else {}

        option_key = option or ""
        # 전체 범위가 기록된 경우 우선 사용
        all_key = f"{option_key}::__all__"
        ts_str: Optional[str] = audits.get(all_key)
        partial = False

        # 개별 로케이션 실사 시간이 있으면 그것도 집계
        per_location_dates: Dict[str, date] = {}
        if not ts_str:
            for loc in locations:
                loc_key = f"{option_key}::{loc}"
                if loc_key in audits:
                    try:
                        per_location_dates[loc] = datetime.fromisoformat(str(audits[loc_key])).date()
                    except Exception:
                        continue
            if per_location_dates:
                ts = max(per_location_dates.values())
                ts_str = ts.isoformat()
                partial = len(per_location_dates) < len(locations)

        if not ts_str:
            return "미실사", "audit_old", False

        try:
            ts = datetime.fromisoformat(str(ts_str)).date()
        except Exception:
            return "미실사", "audit_old", partial

        today = date.today()
        days = max(0, (today - ts).days)
        if days < 30:
            label = f"{days}일"
        else:
            label = f"{days / 30:.1f}개월"

        if days <= 90:
            tag = "audit_recent"
        elif days <= 180:
            tag = "audit_mid"
        else:
            tag = "audit_old"
        return label, tag, partial

    def _record_last_audit(self, item: str, option: str, scope: str) -> None:
        metadata = self.data.setdefault("item_metadata", {})
        info = metadata.setdefault(item, {})
        audits = info.setdefault("last_audit", {})
        option_key = option or ""
        audits[f"{option_key}::{scope}"] = datetime.now().date().isoformat()

    def _record_last_audit_for_row(self, row: Dict[str, object], scope: Optional[str] = None) -> None:
        target_scope = scope or row.get("audit_scope") or "__all__"
        self._record_last_audit(row.get("item", ""), row.get("option", ""), target_scope)

    def refresh_stock(self) -> None:
        rows = self._generate_stock_rows()
        self.stock_rows = []
        self.stock_row_lookup = {row["id"]: row for row in rows}
        self.checked_stock_ids &= set(self.stock_row_lookup)
        for row_id in self.stock_tree.get_children():
            self.stock_tree.delete(row_id)
        for idx, row in enumerate(rows):
            audit_label = row.get("audit_label", "미실사")
            if row.get("audit_partial"):
                audit_label = f"{audit_label} ⚠"
            category_label = self._category_label(row.get("category", "album"))
            values = (
                "☑" if row["id"] in self.checked_stock_ids else "☐",
                audit_label,
                category_label,
                row["artist"],
                row["item"],
                row["option"] or "-",
                self._format_quantity(row["opening"]),
                self._format_quantity(row["in_total"]),
                self._format_quantity(row["out_total"]),
                self._format_quantity(row["qty"]),
                "로케이션 확인",
            )
            tags = {"even" if idx % 2 == 0 else "odd"}
            if row["qty"] <= 0:
                tags.add("negative")
            if row.get("audit_tag"):
                tags.add(row["audit_tag"])
            if row.get("audit_partial"):
                tags.add("audit_partial")
            self.stock_tree.insert("", tk.END, iid=row["id"], values=values, tags=list(tags))
            self.stock_rows.append(
                (
                    category_label,
                    row["artist"],
                    row["item"],
                    row["option"],
                    row["opening"],
                    row["in_total"],
                    row["out_total"],
                    row["qty"],
                    audit_label,
                    row["location_display"],
                )
            )

    def _on_stock_click(self, event) -> Optional[str]:
        region = self.stock_tree.identify("region", event.x, event.y)
        if region != "cell":
            return None
        column = self.stock_tree.identify_column(event.x)
        row_id = self.stock_tree.identify_row(event.y)
        try:
            col_index = int(column.lstrip("#")) - 1
        except ValueError:
            return None
        display_cols = list(self.stock_tree["displaycolumns"])
        if 0 <= col_index < len(display_cols):
            col_name = display_cols[col_index]
            if col_name == "select" and row_id:
                self._toggle_stock_checkbox(row_id)
                return "break"
            if col_name == "location" and row_id:
                self.open_location_overview(row_id)
                return "break"
        return None

    def _on_stock_motion(self, event) -> None:
        region = self.stock_tree.identify("region", event.x, event.y)
        column = self.stock_tree.identify_column(event.x)
        row_id = self.stock_tree.identify_row(event.y)
        if region == "cell" and row_id:
            try:
                col_index = int(column.lstrip("#")) - 1
            except ValueError:
                self._clear_location_hover()
                self.stock_tree.configure(cursor="")
                return
            display_cols = list(self.stock_tree["displaycolumns"])
            if 0 <= col_index < len(display_cols) and display_cols[col_index] == "location":
                if self._hover_location_row != row_id:
                    self._clear_location_hover()
                    self._hover_location_row = row_id
                    tags = set(self.stock_tree.item(row_id, "tags"))
                    tags.add("location_hover")
                    self.stock_tree.item(row_id, tags=list(tags))
                self.stock_tree.configure(cursor="hand2")
                return
        self._clear_location_hover()
        self.stock_tree.configure(cursor="")

    def _clear_location_hover(self) -> None:
        if not getattr(self, "_hover_location_row", None):
            return
        row_id = self._hover_location_row
        tags = set(self.stock_tree.item(row_id, "tags"))
        if "location_hover" in tags:
            tags.remove("location_hover")
            self.stock_tree.item(row_id, tags=list(tags))
        self._hover_location_row = None

    def _choose_locations(self, row: Dict[str, object], *, multiple: bool) -> Optional[object]:
        locations = sorted((row.get("locations") or {}).keys())
        if not locations:
            messagebox.showinfo("안내", "선택한 품목의 로케이션 정보를 찾을 수 없습니다.")
            return None
        if len(locations) == 1:
            return locations if multiple else locations[0]

        result: Optional[object] = None
        top = tk.Toplevel(self.root)
        top.title("로케이션 선택")
        top.transient(self.root)
        top.grab_set()
        ttk.Label(top, text="작업할 로케이션을 선택하세요.", padding=12).pack(anchor="w")
        container = ttk.Frame(top, padding=(12, 0, 12, 12))
        container.pack(fill=tk.BOTH, expand=True)
        if multiple:
            vars: Dict[str, tk.BooleanVar] = {}
            for loc in locations:
                var = tk.BooleanVar(value=True)
                vars[loc] = var
                ttk.Checkbutton(container, text=loc, variable=var).pack(anchor="w")

            def on_ok():
                selected = [loc for loc, var in vars.items() if var.get()]
                nonlocal result
                result = selected
                top.destroy()

        else:
            choice = tk.StringVar(value=locations[0])
            for loc in locations:
                ttk.Radiobutton(container, text=loc, value=loc, variable=choice).pack(anchor="w")

            def on_ok():
                nonlocal result
                result = choice.get()
                top.destroy()

        button_row = ttk.Frame(top, padding=12)
        button_row.pack(fill=tk.X)
        ttk.Button(button_row, text="확인", command=on_ok).pack(side=tk.LEFT)
        ttk.Button(button_row, text="취소", command=top.destroy).pack(side=tk.LEFT, padx=(8, 0))
        top.wait_window()
        return result

    def open_location_overview(self, row_id: str) -> None:
        row = self.stock_row_lookup.get(row_id)
        if not row:
            messagebox.showinfo("안내", "선택한 재고 정보를 찾을 수 없습니다.")
            return
        stock = self.data.get("stock", {})
        option_map = stock.get(row["item"], {}).get(row["option"], {})
        if not option_map:
            messagebox.showinfo("안내", "해당 품목의 로케이션 정보를 찾을 수 없습니다.")
            return
        total_qty = sum(option_map.values())
        top = tk.Toplevel(self.root)
        top.title("로케이션 확인")
        top.transient(self.root)
        top.grab_set()
        top.geometry("640x480")
        top.minsize(560, 380)

        wrapper = ttk.Frame(top, padding=16)
        wrapper.pack(fill=tk.BOTH, expand=True)

        header = ttk.LabelFrame(wrapper, text="선택한 재고", padding=12)
        header.pack(fill=tk.X)
        ttk.Label(header, text=f"아티스트: {row['artist']}").grid(row=0, column=0, sticky=tk.W, padx=(0, 12))
        ttk.Label(header, text=f"앨범/버전: {row['item']}").grid(row=0, column=1, sticky=tk.W)
        ttk.Label(header, text=f"옵션: {row['option'] or '-'}").grid(row=1, column=0, sticky=tk.W, pady=(8, 0))
        ttk.Label(header, text=f"현재고: {self._format_quantity(total_qty)}").grid(
            row=1, column=1, sticky=tk.W, pady=(8, 0)
        )
        ttk.Label(header, text=f"로케이션 수: {len(option_map)}곳").grid(row=2, column=0, sticky=tk.W, pady=(8, 0))

        list_frame = ttk.LabelFrame(wrapper, text="로케이션 목록", padding=12)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(12, 0))

        columns = ("location", "qty")
        tree = ttk.Treeview(list_frame, columns=columns, show="headings", selectmode="browse")
        tree.heading("location", text="로케이션")
        tree.heading("qty", text="수량")
        tree.column("location", width=260, anchor=tk.CENTER)
        tree.column("qty", width=100, anchor=tk.CENTER)
        tree.tag_configure("negative", foreground="#b91c1c")

        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        for location, qty in sorted(option_map.items()):
            tags = ("negative",) if qty <= 0 else ()
            tree.insert("", tk.END, values=(location, self._format_quantity(qty)), tags=tags)

        def copy_selected(event=None):
            selection = tree.selection()
            if not selection:
                return
            loc_val = tree.item(selection[0], "values")[0]
            if not self._show_location_on_map(loc_val):
                self._handle_location_button(loc_val)

        tree.bind("<Double-1>", copy_selected)

        helper = ttk.Frame(wrapper)
        helper.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(
            helper,
            text="로케이션을 더블클릭하면 지도에서 확인하거나 값을 복사할 수 있습니다.",
            anchor=tk.W,
        ).pack(side=tk.LEFT)
        ttk.Button(helper, text="닫기", command=top.destroy).pack(side=tk.RIGHT)

    def _handle_location_button(self, location: str) -> None:
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(location)
            self.set_status(f"로케이션 '{location}'을 복사했습니다.")
        except tk.TclError:
            self.set_status("로케이션 정보를 처리할 수 없습니다.")

    def _show_location_on_map(self, location: str) -> bool:
        entries = [entry for entry in load_location_entries(LOCATION_MAP_FILE) if entry.get("location") == location]
        if not entries:
            return False
        try:
            LocationMapViewer(self.root, entries)
            return True
        except Exception as exc:  # pragma: no cover - runtime UI
            messagebox.showerror("지도 표시 실패", f"로케이션 지도를 여는 중 오류가 발생했습니다: {exc}")
            return False

    def _toggle_stock_checkbox(self, row_id: str) -> None:
        if row_id in self.checked_stock_ids:
            self.checked_stock_ids.remove(row_id)
        else:
            self.checked_stock_ids.add(row_id)
        if row_id in self.stock_row_lookup:
            row = self.stock_row_lookup[row_id]
            values = (
                "☑" if row_id in self.checked_stock_ids else "☐",
                row["artist"],
                row["item"],
                row["option"] or "-",
                self._format_quantity(row["opening"]),
                self._format_quantity(row["in_total"]),
                self._format_quantity(row["out_total"]),
                self._format_quantity(row["qty"]),
                "로케이션 확인",
            )
            self.stock_tree.item(row_id, values=values)

    def _generate_stock_rows(self) -> List[Dict[str, object]]:
        stock = self.data.get("stock", {})
        metadata = self.data.get("item_metadata", {})
        selection = self.artist_filter_var.get()
        category_choice = getattr(self, "category_filter_var", tk.StringVar(value="전체")).get()
        location_selection = getattr(self, "location_filter_var", tk.StringVar(value="전체")).get()
        filtered_stock = stock
        if selection and selection != "전체":
            filtered_stock = filter_stock_by_artist(stock, metadata, selection)

        category_key = self._normalize_category(category_choice) if category_choice and category_choice != "전체" else None

        location_pool: Set[str] = set(self.settings.get("location_presets", []))
        for item_locations in stock.values():
            for option_locations in item_locations.values():
                location_pool.update(option_locations.keys())
        self._refresh_location_filter_options(sorted(location_pool))
        location_selection = getattr(self, "location_filter_var", tk.StringVar(value="전체")).get()

        period = self.data.get("current_period")
        opening = {}
        if period:
            opening = self.data.get("periods", {}).get(period, {}).get("opening_stock", {})
        activity = self._calculate_period_activity(period, location_selection)
        rows: List[Dict[str, object]] = []
        for item in sorted(filtered_stock):
            category_value = self._normalize_category(metadata.get(item, {}).get("category", "album"))
            if category_key and category_value != category_key:
                continue
            for option in sorted(filtered_stock[item]):
                all_locations = filtered_stock[item][option]
                if location_selection and location_selection != "전체":
                    locations = {location_selection: all_locations.get(location_selection, 0)}
                    if sum(locations.values()) == 0:
                        continue
                else:
                    locations = all_locations

                total_qty = sum(locations.values())
                opening_map = opening.get(item, {}).get(option, {})
                if location_selection and location_selection != "전체":
                    opening_total = opening_map.get(location_selection, 0)
                    location_display = self._format_location_detail({location_selection: locations.get(location_selection, 0)})
                    audit_scope = location_selection
                else:
                    opening_total = sum(opening_map.values())
                    location_display = self._format_location_detail(locations)
                    audit_scope = "__all__"

                metrics = activity.get((item, option), {"in": 0, "out": 0})
                audit_label, audit_tag, audit_partial = self._audit_status_for_row(item, option, locations)
                rows.append(
                    {
                        "id": f"{item}::{option}"
                        if location_selection == "전체" or not location_selection
                        else f"{item}::{option}::{location_selection}",
                        "category": category_value,
                        "artist": metadata.get(item, {}).get("artist") or "-",
                        "item": item,
                        "option": option,
                        "location_display": location_display,
                        "location_scope": location_selection if location_selection != "전체" else "",
                        "opening": opening_total,
                        "in_total": metrics.get("in", 0),
                        "out_total": metrics.get("out", 0),
                        "qty": total_qty,
                        "locations": deepcopy(all_locations),
                        "audit_label": audit_label,
                        "audit_tag": audit_tag,
                        "audit_scope": audit_scope,
                        "audit_partial": audit_partial and len(locations) > 1,
                    }
                )
        rows.sort(
            key=lambda r: (
                self._category_label(r.get("category")),
                r.get("artist") or "",
                r.get("item") or "",
                r.get("option") or "",
                r.get("location_display") or "",
            )
        )
        return rows

    def _calculate_period_activity(self, period: Optional[str], location_filter: str = "전체") -> Dict[Tuple[str, str], Dict[str, int]]:
        history = self.data.get("history", [])
        cache_key = (period or "", location_filter or "전체", len(history), self.data.get("last_updated"))
        if cache_key in self._activity_cache:
            return self._activity_cache[cache_key]

        activity: Dict[Tuple[str, str], Dict[str, int]] = {}
        if not period:
            return activity
        for entry in history:
            if entry.get("period") != period:
                continue
            if location_filter and location_filter != "전체":
                if entry.get("location") != location_filter:
                    continue
            key = (entry.get("item"), entry.get("option", ""))
            if not key[0]:
                continue
            tracker = activity.setdefault(key, {"in": 0, "out": 0})
            tx_type = entry.get("type")
            if tx_type not in ("in", "out"):
                continue
            tracker[tx_type] = tracker.get(tx_type, 0) + entry.get("quantity", 0)
        self._activity_cache[cache_key] = activity
        return activity

    def _default_history_range(self) -> Tuple[str, str]:
        today = date.today()
        start = (today - timedelta(days=29)).isoformat()
        return start, today.isoformat()

    def _normalize_history_dates(
        self, start_day: Optional[str], end_day: Optional[str]
    ) -> Tuple[str, str]:
        if not start_day and not end_day:
            return self._default_history_range()
        if start_day and not end_day:
            return start_day, start_day
        if end_day and not start_day:
            return end_day, end_day
        if start_day and end_day and start_day > end_day:
            return end_day, start_day
        return start_day or "", end_day or ""

    def _parse_day_input(self, raw: str) -> date:
        cleaned = raw.strip()
        if not cleaned:
            return date.today()
        try:
            return datetime.strptime(cleaned, "%Y-%m-%d").date()
        except ValueError:
            pass
        normalized = cleaned.replace(".", "/")
        for sep in ("/", "-"):
            if sep in normalized:
                parts = [p for p in normalized.split(sep) if p]
                if len(parts) == 2:
                    try:
                        month = int(parts[0])
                        day = int(parts[1])
                        return date(datetime.now().year, month, day)
                    except ValueError:
                        continue
        raise ValueError("날짜 형식을 인식할 수 없습니다. 예: 2024-05-01 또는 11/20")

    def open_tx_calendar(self) -> None:
        existing = self.tx_date_var.get().strip()
        initial = None
        if existing:
            try:
                initial = datetime.strptime(existing, "%Y-%m-%d").date()
            except ValueError:
                initial = None
        CalendarPopup(self.root, lambda d: self.tx_date_var.set(d.isoformat()), initial_date=initial)

    def open_history_calendar(self, target_var: tk.StringVar) -> None:
        existing = target_var.get().strip()
        initial = None
        if existing:
            try:
                initial = datetime.strptime(existing, "%Y-%m-%d").date()
            except ValueError:
                initial = None
        CalendarPopup(self.root, lambda d: self._on_history_date(target_var, d), initial_date=initial)

    def _on_history_date(self, target_var: tk.StringVar, selected: date) -> None:
        target_var.set(selected.isoformat())
        self.search_history("in", triggered_by_calendar=True)
        self.search_history("out", triggered_by_calendar=True)

    def fill_transaction_from_stock(self) -> None:
        selection = self.stock_tree.selection()
        if not selection:
            return
        row = self.stock_row_lookup.get(selection[0])
        if not row:
            return
        self.artist_var.set(row.get("artist") if row.get("artist") != "-" else "")
        self.category_var.set(self._category_label(row.get("category", "album")))
        self.item_var.set(row.get("item", ""))
        self.option_var.set(row.get("option") if row.get("option") != "-" else "")
        locations = row.get("locations", {})
        filtered_location = getattr(self, "location_filter_var", tk.StringVar(value="전체")).get()
        if filtered_location and filtered_location != "전체" and filtered_location in locations:
            location_value = filtered_location
        else:
            location_value = list(locations)[0] if len(locations) == 1 else ""
        self.location_var.set(location_value)
        self.quantity_var.set("")
        self.tx_date_var.set(date.today().isoformat())
        self.set_status("입/출고 기록 폼을 선택한 재고 정보로 채웠습니다.")

    def edit_selected_stock(self) -> None:
        selection = self.stock_tree.selection()
        if not selection:
            messagebox.showinfo("안내", "수정할 재고를 선택해 주세요.")
            return
        row_id = selection[0]
        row = self.stock_row_lookup.get(row_id)
        if not row:
            messagebox.showinfo("안내", "선택한 재고 정보를 찾을 수 없습니다.")
            return
        selected_location = self._choose_locations(row, multiple=False)
        if not selected_location:
            return
        dialog = StockEditDialog(
            self.root,
            item=row["item"],
            location=selected_location,
            quantity=row["locations"].get(selected_location, 0),
            artist=row["artist"],
            option=row["option"],
            category=row.get("category", "album"),
        )
        self.root.wait_window(dialog)
        if not dialog.result:
            return
        self._apply_stock_edit(row["item"], row["option"], selected_location, dialog.result)

    def delete_selected_stock(self) -> None:
        if not self.checked_stock_ids:
            messagebox.showinfo("안내", "삭제할 재고에 체크해 주세요.")
            return

        targets = [self.stock_row_lookup[row_id] for row_id in self.checked_stock_ids if row_id in self.stock_row_lookup]
        if not targets:
            messagebox.showinfo("안내", "삭제 대상 재고를 찾지 못했습니다.")
            return

        deletions: List[Tuple[Dict[str, object], List[str]]] = []
        for row in targets:
            chosen = self._choose_locations(row, multiple=True)
            if chosen is None:
                return
            deletions.append((row, chosen if isinstance(chosen, list) else [chosen]))

        names = ", ".join(
            {
                f"{row['artist']} - {row['item']} {row['option']} ({'/'.join(locs)})"
                for row, locs in deletions
            }
        )
        confirm = messagebox.askyesno("확인", f"선택한 로케이션을 삭제하시겠습니까?\n{names}")
        if not confirm:
            return

        stock = self.data.get("stock", {})
        metadata = self.data.get("item_metadata", {})
        for row, selected_locations in deletions:
            option_map = stock.get(row["item"], {})
            locations = option_map.get(row["option"], {})
            for loc in selected_locations:
                locations.pop(loc, None)
            if not locations:
                option_map.pop(row["option"], None)
            if not option_map:
                stock.pop(row["item"], None)
                metadata.pop(row["item"], None)
        self._save_async()
        self.checked_stock_ids.clear()
        self._refresh_artist_options()
        self.refresh_stock()
        self.set_status("선택한 재고를 삭제했습니다.")
        self._log_user_action("선택 재고 삭제", persist=False)

    def open_stock_audit(self) -> None:
        if self.audit_window and self.audit_window.winfo_exists():
            self.audit_window.lift()
            self.audit_window.focus()
            return

        self.refresh_stock()
        self.audit_counts = {}
        self.audit_entry_map = {}
        self.audit_count_var.set("")
        self.audit_window = tk.Toplevel(self.root)
        self.audit_window.title("재고 실사 모드")
        self.audit_window.geometry("960x540")
        self.audit_window.transient(self.root)
        self.audit_window.grab_set()
        self.audit_window.protocol("WM_DELETE_WINDOW", self._close_audit_window)

        frame = ttk.Frame(self.audit_window, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="현재 재고를 기준으로 실사 수량을 입력하세요. 완료된 항목은 하이라이트됩니다.").pack(
            anchor="w", pady=(0, 8)
        )

        columns = ("artist", "item", "option", "location", "audit", "current", "counted", "status")
        self.audit_tree = ttk.Treeview(frame, columns=columns, show="headings", height=16)
        headings = [
            ("artist", "아티스트", 140),
            ("item", "앨범/버전", 200),
            ("option", "옵션", 120),
            ("location", "로케이션", 200),
            ("audit", "실사 경과", 100),
            ("current", "현재고", 80),
            ("counted", "실사 수량", 90),
            ("status", "상태", 90),
        ]
        for col, title, width in headings:
            self.audit_tree.heading(col, text=title)
            self.audit_tree.column(col, width=width, anchor=tk.CENTER)
        self.audit_tree.tag_configure("audited", background="#ecfeff")
        self.audit_tree.tag_configure("audit_recent", foreground="#166534", background="#ecfdf3")
        self.audit_tree.tag_configure("audit_mid", foreground="#92400e", background="#fffbeb")
        self.audit_tree.tag_configure("audit_old", foreground="#991b1b", background="#fef2f2")
        self.audit_tree.tag_configure("even", background="#f8fafc")
        self.audit_tree.tag_configure("odd", background="#ffffff")
        self.audit_tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        self.audit_tree.bind("<<TreeviewSelect>>", self._on_audit_select)

        yscroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.audit_tree.yview)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.audit_tree.configure(yscrollcommand=yscroll.set)

        control = ttk.Frame(self.audit_window, padding=(12, 0, 12, 12))
        control.pack(fill=tk.X)
        ttk.Label(control, text="선택 품목 실사 수량").pack(side=tk.LEFT)
        ttk.Entry(control, textvariable=self.audit_count_var, width=10).pack(side=tk.LEFT, padx=(6, 8))
        ttk.Button(control, text="선택 품목 실사 완료", command=self._mark_audit_complete).pack(side=tk.LEFT)
        ttk.Button(control, text="실사 초기화", command=self._reset_audit_counts).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(control, text="실사 결과 일괄 적용", command=self._apply_audit_adjustments).pack(side=tk.RIGHT)

        self._populate_audit_tree()

    def _close_audit_window(self) -> None:
        if self.audit_window and self.audit_window.winfo_exists():
            self.audit_window.destroy()
        self.audit_window = None

    def _populate_audit_tree(self) -> None:
        if not hasattr(self, "audit_tree"):
            return
        tree = self.audit_tree
        tree.delete(*tree.get_children())
        self.audit_entry_map = {}
        insert_idx = 0
        for row_id, row in self.stock_row_lookup.items():
            locations = row.get("locations", {})
            # 단일 로케이션 또는 필터링된 경우
            if len(locations) <= 1:
                loc = next(iter(locations)) if locations else ""
                entry_id = row_id if not loc else f"{row_id}::{loc}"
                self._insert_audit_entry(tree, entry_id, row, loc, row.get("qty", 0), insert_idx)
                insert_idx += 1
                continue

            # 복수 로케이션은 각각 별도 행으로 노출
            for loc, qty in sorted(locations.items()):
                entry_id = f"{row_id}::{loc}"
                self._insert_audit_entry(tree, entry_id, row, loc, qty, insert_idx)
                insert_idx += 1

    def _insert_audit_entry(
        self, tree: ttk.Treeview, entry_id: str, row: Dict[str, object], location: str, current_qty: int, row_index: int
    ) -> None:
        counted = self.audit_counts.get(entry_id)
        status = "완료" if counted is not None else "대기"
        display_count = self._format_quantity(counted) if counted is not None else ""
        tags: List[str] = ["even" if row_index % 2 == 0 else "odd"]
        label, tag, _ = self._audit_status_for_row(
            row.get("item", ""),
            row.get("option", ""),
            {location: current_qty} if location else row.get("locations", {}),
        )
        if counted is not None:
            tags.append("audited")
        if tag:
            tags.append(str(tag))
        location_display = (
            self._format_location_detail({location: current_qty})
            if location
            else self._format_location_detail(row.get("locations", {}))
        )
        tree.insert(
            "",
            tk.END,
            iid=entry_id,
            values=(
                row.get("artist"),
                row.get("item"),
                row.get("option") or "-",
                location_display,
                label,
                self._format_quantity(current_qty),
                display_count,
                status,
            ),
            tags=tags,
        )
        self.audit_entry_map[entry_id] = {"row_id": row.get("id"), "row": row, "location": location}

    def _on_audit_select(self, event=None) -> None:
        selection = self.audit_tree.selection()
        if not selection:
            return
        row_id = selection[0]
        existing = self.audit_counts.get(row_id)
        self.audit_count_var.set("" if existing is None else str(existing))

    def _mark_audit_complete(self) -> None:
        selection = self.audit_tree.selection()
        if not selection:
            messagebox.showinfo("안내", "실사 수량을 입력할 품목을 선택해 주세요.")
            return
        try:
            counted = int(self.audit_count_var.get().strip())
        except ValueError:
            messagebox.showerror("오류", "실사 수량은 0 이상의 정수로 입력해 주세요.")
            return
        if counted < 0:
            messagebox.showerror("오류", "실사 수량은 음수가 될 수 없습니다.")
            return
        entry_info = self.audit_entry_map.get(selection[0])
        if not entry_info:
            messagebox.showerror("오류", "선택한 항목 정보를 찾을 수 없습니다.")
            return
        row = entry_info.get("row")
        loc = entry_info.get("location")
        if not row:
            messagebox.showerror("오류", "선택한 재고 정보를 찾을 수 없습니다.")
            return
        current_qty = row.get("locations", {}).get(loc, row.get("qty", 0))
        if counted != current_qty:
            proceed = messagebox.askyesno(
                "확인",
                (
                    f"현재 재고는 {self._format_quantity(current_qty)}개입니다.\n"
                    f"입력한 실사 수량 {self._format_quantity(counted)}개로 기록할까요?"
                ),
            )
            if not proceed:
                self.audit_tree.focus(selection[0])
                self.audit_tree.selection_set(selection[0])
                return
        self.audit_counts[selection[0]] = counted
        self._populate_audit_tree()

    def _reset_audit_counts(self) -> None:
        self.audit_counts = {}
        self.audit_count_var.set("")
        self._populate_audit_tree()

    def _apply_audit_adjustments(self) -> None:
        if not self.audit_counts:
            messagebox.showinfo("안내", "실사 완료된 품목이 없습니다. 먼저 수량을 입력해 주세요.")
            return
        total_items = len(self.audit_entry_map)
        if len(self.audit_counts) < total_items:
            proceed = messagebox.askyesno(
                "확인",
                f"전체 {total_items}개 중 {len(self.audit_counts)}개만 실사되었습니다. 계속 적용할까요?",
            )
            if not proceed:
                return

        adjustments = 0
        audit_updates = False
        for entry_id, counted in self.audit_counts.items():
            entry_info = self.audit_entry_map.get(entry_id, {})
            row_id = entry_info.get("row_id")
            row = self.stock_row_lookup.get(row_id) if row_id else None
            if not row:
                continue
            location_scope = entry_info.get("location") or row.get("audit_scope") or "__all__"
            current_qty = row.get("locations", {}).get(location_scope, row.get("qty", 0))
            delta = counted - current_qty
            self._record_last_audit(row.get("item", ""), row.get("option", ""), location_scope)
            audit_updates = True
            if delta == 0:
                continue
            desc = "물류실사 실재고 증가분" if delta > 0 else "물류실사 실재고 감소분"
            meta_artist = row.get("artist")
            if not meta_artist or meta_artist == "-":
                meta_artist = self.data.get("item_metadata", {}).get(row["item"], {}).get("artist") or "미분류"
            location_for_tx = location_scope or ""
            if not location_for_tx:
                location_for_tx = sorted((row.get("locations") or {}).keys())[0:1]
                location_for_tx = location_for_tx[0] if location_for_tx else "실사조정"
            tx = Transaction(
                type="in" if delta > 0 else "out",
                artist=meta_artist,
                item=row["item"],
                category=row.get("category", "album"),
                option=row.get("option", ""),
                location=location_for_tx,
                quantity=abs(delta),
                timestamp=datetime.now(),
                actor=self._current_actor(),
                description=desc,
            )
            try:
                record_transaction(self.data, tx, allow_negative=True)
            except ValueError as exc:
                messagebox.showerror("오류", str(exc))
                return
            adjustments += 1

        if adjustments or audit_updates:
            self._save_async()
            self.refresh_stock()
            self.search_history("in", triggered_by_calendar=True)
            self.search_history("out", triggered_by_calendar=True)
            if adjustments:
                self.set_status(f"실사 결과를 적용했습니다. 조정 {adjustments}건")
            else:
                self.set_status("실사 시점을 업데이트했습니다.")
            self._log_user_action(
                f"실사 결과 적용 - 조정 {adjustments}건, 실사 {len(self.audit_counts)}건",
                persist=False,
            )
        else:
            self.set_status("실사 결과와 현재 재고가 동일합니다.")
        self._close_audit_window()

    def _apply_stock_edit(self, old_item: str, old_option: str, old_location: str, new_values: Dict[str, object]) -> None:
        metadata = self.data.setdefault("item_metadata", {})
        stock = self.data.setdefault("stock", {})

        new_item = str(new_values.get("item", old_item)).strip()
        new_artist = str(new_values["artist"])
        new_category = self._normalize_category(str(new_values.get("category", "album")))
        new_option = str(new_values.get("option", ""))
        new_location = str(new_values["location"])
        new_qty = int(new_values["quantity"])
        original_qty = stock.get(old_item, {}).get(old_option or "", {}).get(old_location, 0)

        delta = new_qty - original_qty
        allow_negative = False
        if delta < 0 and abs(delta) > original_qty:
            proceed = messagebox.askyesno(
                "확인",
                "현재 수량보다 적은 값으로 수정됩니다. 음수 재고로 자체수정 기록을 남길까요?",
            )
            if not proceed:
                return
            allow_negative = True
        if new_qty < 0:
            proceed = messagebox.askyesno("확인", "음수 재고로 수정하시겠습니까?")
            if not proceed:
                return
            allow_negative = True

        existing_meta = metadata.get(new_item)
        if new_item != old_item and existing_meta and existing_meta.get("artist") not in (None, new_artist):
            messagebox.showerror(
                "오류",
                f"이미 {existing_meta.get('artist')} 아티스트로 등록된 품목명이 존재합니다. 같은 아티스트명으로 변경해 주세요.",
            )
            return

        if new_item != old_item:
            moved_meta = metadata.pop(old_item, {})
            merged_meta = {**moved_meta, "artist": new_artist, "category": new_category}
            if new_option:
                merged_meta["option"] = new_option
            metadata[new_item] = merged_meta
        else:
            meta_entry = metadata.setdefault(old_item, {})
            meta_entry["artist"] = new_artist
            meta_entry["category"] = new_category
            if new_option:
                meta_entry["option"] = new_option

        option_key_old = old_option or ""
        option_key_new = new_option or ""
        source_options = stock.get(old_item, {})
        if new_item != old_item:
            source_options = stock.pop(old_item, {})

        target_options = stock.setdefault(new_item, {})
        for opt_key, locations in source_options.items():
            if new_item != old_item or opt_key != option_key_old:
                target_options.setdefault(opt_key, {}).update(locations)

        # 원래 위치 제거 후 새 위치에 재배치
        original_locations = target_options.setdefault(option_key_old, {})
        if original_locations is source_options.get(option_key_old):
            pass
        original_locations.pop(old_location, None)
        if not original_locations and option_key_old in target_options and option_key_new != option_key_old:
            target_options.pop(option_key_old, None)

        new_locations = target_options.setdefault(option_key_new, {})
        new_locations[new_location] = original_qty

        if delta:
            tx = Transaction(
                type="in" if delta > 0 else "out",
                artist=new_artist,
                item=new_item,
                option=new_option,
                location=new_location,
                quantity=abs(delta),
                timestamp=datetime.now(),
                actor=self._current_actor(),
                description="자체수정",
            )
            try:
                record_transaction(self.data, tx, allow_negative=allow_negative)
            except ValueError as exc:
                messagebox.showerror("오류", str(exc))
                return
        self._save_async()
        self._refresh_artist_options()
        self.refresh_stock()
        self.search_history("in", triggered_by_calendar=True)
        self.search_history("out", triggered_by_calendar=True)
        self.set_status("재고를 수정했습니다.")
        self._log_user_action(
            f"재고 수정 - {new_item} / {new_option or '-'} @ {new_location} {new_qty}개",
            persist=False,
        )

    def export_stock(self) -> None:
        if not self.stock_rows:
            messagebox.showinfo("안내", "저장할 현재 재고가 없습니다.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All Files", "*.*")],
            title="현재 재고 엑셀 저장",
        )
        if not path:
            return
        try:
            per_location_rows: List[Tuple[str, str, str, str, str, int]] = []
            for row in self.stock_row_lookup.values():
                for loc, qty in sorted((row.get("locations") or {}).items()):
                    per_location_rows.append(
                        (
                            self._category_label(row.get("category", "album")),
                            row.get("artist"),
                            row.get("item"),
                            row.get("option"),
                            loc,
                            qty,
                        )
                    )
            export_stock_rows_to_xlsx(self.stock_rows, per_location_rows, path)
        except RuntimeError as exc:
            messagebox.showerror("오류", str(exc))
            return
        messagebox.showinfo("완료", f"현재 재고를 엑셀로 저장했습니다: {path}")

    # ---------------------------------------------------------------- transactions
    def submit_transaction(self, tx_type: str) -> None:
        try:
            item = self.item_var.get().strip()
            option = self.option_var.get().strip()
            location = self.location_var.get().strip()
            quantity_str = self.quantity_var.get().strip()
            artist_input = self.artist_var.get().strip() or None
            category_input = self.category_var.get().strip() or "앨범"
            record_day = self.tx_date_var.get().strip()
            description = self.tx_detail_var.get().strip()
            event_mode = self.event_var.get()
            if not item or not location or not quantity_str or not (artist_input or ""):
                raise ValueError("아티스트, 앨범/버전, 로케이션, 수량을 모두 입력해 주세요.")
            if tx_type == "out" and not description:
                raise ValueError("출고 시 상세내용은 필수입니다.")
            quantity = int(quantity_str)
            tx_date = self._parse_day_input(record_day)
            now_time = datetime.now().time().replace(microsecond=0)
            timestamp = datetime.combine(tx_date, now_time)
            artist = determine_artist(self.data, item, artist_input)
            category = determine_category(self.data, item, category_input)
            allow_negative = False
            event_id = ""
            merge_index: Optional[int] = None
            if tx_type == "out":
                option_key = option or ""
                available = (
                    self.data.get("stock", {})
                    .get(item, {})
                    .get(option_key, {})
                    .get(location, 0)
                )
                if quantity > available:
                    proceed = messagebox.askyesno(
                        "확인",
                        f"현재 재고({available}개)보다 많은 출고입니다. 음수 재고로 기록하시겠습니까?",
                    )
                    if not proceed:
                        self.set_status("출고를 취소했습니다.", error=True)
                        return
                    allow_negative = True
                if event_mode:
                    event_id, merge_index = self._determine_event_session(artist, item, option)
            elif event_mode:
                event_id = self._pick_event_for_return(artist, item, option)
                if not event_id:
                    raise ValueError("입고할 이벤트 출고 내역을 선택해 주세요.")
            tx = Transaction(
                type="in" if tx_type == "in" else "out",
                artist=artist,
                item=item,
                category=category,
                option=option,
                location=location,
                quantity=quantity,
                timestamp=timestamp,
                actor=self._current_actor(),
                description=description,
                event=event_mode,
                event_id=event_id,
                event_open=event_mode and tx_type == "out",
            )
            record_transaction(self.data, tx, allow_negative=allow_negative)
            if event_mode and tx_type == "out" and merge_index is not None:
                self._merge_event_out(merge_index, quantity)
            elif event_mode and tx_type == "in" and event_id:
                self._close_event_out(event_id)
            self._save_async()
            self._log_user_action(
                f"{self._current_actor()} {tx.type.upper()} - {item} / {option or '-'} @ {location} {quantity}개",
                persist=False,
            )
        except ValueError as exc:
            messagebox.showerror("오류", str(exc))
            self.set_status(str(exc), error=True)
            return
        self.set_status("기록이 저장되었습니다.")
        self._refresh_artist_options()
        self.refresh_stock()
        self.search_history("in", triggered_by_calendar=True)
        self.search_history("out", triggered_by_calendar=True)
        self.item_var.set("")
        self.option_var.set("")
        self.location_var.set("")
        self.quantity_var.set("1")
        self.tx_detail_var.set("")
        self.tx_date_var.set(date.today().isoformat())

    # ---------------------------------------------------------------- history
    def search_history(
        self, tx_type: str, triggered_by_calendar: bool = False, event_only: bool = False, event_open_only: bool = False
    ) -> None:
        start_day_raw = self.history_start_var.get().strip() or None
        end_day_raw = self.history_end_var.get().strip() or None
        artist_value = self.history_artist_var.get().strip()
        artist = None if artist_value in ("", "전체") else artist_value
        start_day, end_day = self._normalize_history_dates(start_day_raw, end_day_raw)
        self.history_start_var.set(start_day)
        self.history_end_var.set(end_day)
        filtered, indices = self._filter_history_with_index(
            tx_type=tx_type,
            start_day=start_day,
            end_day=end_day,
            artist=artist,
            event_only=event_only,
            event_open_only=event_open_only,
        )
        self.history_cache[tx_type] = filtered
        self.history_indices[tx_type] = indices
        self.current_history_type = tx_type
        self.history_event_filter = bool(event_only and event_open_only)
        if event_only:
            caption = "이벤트 출고(미해결)"
        else:
            caption = "입고 결과" if tx_type == "in" else "출고 결과"
        self.history_caption.set(f"{caption} ({start_day} ~ {end_day})")
        self._update_history_tree(filtered)
        if not triggered_by_calendar:
            self.set_status(f"{('입고' if tx_type == 'in' else '출고')} 검색 결과 {len(filtered)}건")

    def _refresh_current_history(self) -> None:
        if self.current_history_type == "out" and self.history_event_filter:
            self.search_history("out", event_only=True, event_open_only=True, triggered_by_calendar=True)
        else:
            self.search_history(self.current_history_type or "in", triggered_by_calendar=True)

    def _filter_history_with_index(
        self,
        *,
        tx_type: str,
        start_day: Optional[str],
        end_day: Optional[str],
        artist: Optional[str],
        event_only: bool = False,
        event_open_only: bool = False,
    ) -> Tuple[List[Dict], List[int]]:
        start = datetime.fromisoformat(start_day).date() if start_day else None
        end = datetime.fromisoformat(end_day).date() if end_day else None
        results: List[Dict] = []
        indices: List[int] = []
        for idx, entry in enumerate(self.data.get("history", [])):
            if entry.get("type") != tx_type:
                continue
            if event_only and not entry.get("event"):
                continue
            if event_open_only and not entry.get("event_open", False):
                continue
            entry_day_str = entry.get("day")
            if start or end:
                if not entry_day_str:
                    continue
                try:
                    entry_day = datetime.fromisoformat(entry_day_str).date()
                except ValueError:
                    continue
                if start and entry_day < start:
                    continue
                if end and entry_day > end:
                    continue
            if artist and entry.get("artist") != artist:
                continue
            results.append(entry)
            indices.append(idx)
        return results, indices

    def _update_history_tree(self, entries: List[Dict]) -> None:
        tree = self.history_tree
        tree.delete(*tree.get_children())
        if not entries:
            tree.insert(
                "",
                tk.END,
                iid="empty",
                values=("-", "-", "-", "-", "-", "-", "-", "조건에 해당하는 내역이 없습니다."),
            )
            return
        for idx, entry in enumerate(entries):
            category_label = self._category_label(entry.get("category", "album"))
            values = (
                entry.get("day"),
                entry.get("artist"),
                category_label,
                entry.get("item"),
                entry.get("option", ""),
                entry.get("location"),
                self._format_quantity(entry.get("quantity", 0)),
                entry.get("description", ""),
            )
            tags = ["even" if idx % 2 == 0 else "odd"]
            if entry.get("event") and entry.get("type") == "out" and entry.get("event_open", False):
                tags.append("event_out")
            tree.insert("", tk.END, iid=str(idx), values=values, tags=tags)

    def _determine_event_session(self, artist: str, item: str, option: str) -> Tuple[str, Optional[int]]:
        key = (artist, item, option or "", self._normalize_category(self.category_var.get() if hasattr(self, "category_var") else "album"))
        history = self.data.get("history", [])
        latest_event = None
        latest_index = None
        for idx in reversed(range(len(history))):
            entry = history[idx]
            if (
                entry.get("event")
                and entry.get("type") == "out"
                and entry.get("event_open", False)
                and (entry.get("artist"), entry.get("item"), entry.get("option", ""), normalize_category(entry.get("category", "album"))) == key
            ):
                latest_event = entry
                latest_index = idx
                break
        new_event_id = datetime.now().strftime("%Y%m%d%H%M%S")
        if not latest_event:
            return new_event_id, None
        reuse = messagebox.askyesno(
            "이벤트 확인",
            "동일 이벤트 출고로 합산하시겠습니까?\n'예'를 누르면 기존 이벤트에 수량이 합산됩니다.",
        )
        if reuse:
            return latest_event.get("event_id") or new_event_id, latest_index
        return new_event_id, None

    def _pick_event_for_return(self, artist: str, item: str, option: str) -> str:
        key = (
            artist,
            item,
            option or "",
            self._normalize_category(self.category_var.get() if hasattr(self, "category_var") else "album"),
        )
        history = self.data.get("history", [])
        open_events: List[Tuple[int, Dict]] = []
        for idx, entry in enumerate(history):
            if (
                entry.get("event")
                and entry.get("type") == "out"
                and entry.get("event_open", False)
                and (
                    entry.get("artist"),
                    entry.get("item"),
                    entry.get("option", ""),
                    normalize_category(entry.get("category", "album")),
                )
                == key
            ):
                open_events.append((idx, entry))
        if not open_events:
            return ""
        if len(open_events) == 1:
            return open_events[0][1].get("event_id", "")
        dialog = EventReturnDialog(self.root, open_events, self._format_quantity)
        self.root.wait_window(dialog)
        if dialog.selection is None:
            return ""
        return open_events[dialog.selection][1].get("event_id", "")

    def _merge_event_out(self, target_index: int, added_qty: int) -> None:
        history = self.data.get("history", [])
        if not history:
            return
        new_entry = history.pop()
        if target_index >= len(history):
            history.append(new_entry)
            return
        target = history[target_index]
        target["quantity"] = int(target.get("quantity", 0)) + added_qty
        target["timestamp"] = new_entry.get("timestamp", target.get("timestamp"))
        target["day"] = new_entry.get("day", target.get("day"))
        target["period"] = new_entry.get("period", target.get("period"))
        target["year"] = new_entry.get("year", target.get("year"))
        target["event"] = True
        target["event_open"] = True
        target.setdefault("event_id", new_entry.get("event_id", ""))

    def _close_event_out(self, event_id: str) -> None:
        if not event_id:
            return
        for entry in self.data.get("history", []):
            if entry.get("event") and entry.get("event_id") == event_id and entry.get("type") == "out":
                entry["event_open"] = False

    def _reopen_event(self, event_id: str) -> None:
        if not event_id:
            return
        for entry in self.data.get("history", []):
            if entry.get("event") and entry.get("event_id") == event_id and entry.get("type") == "out":
                entry["event_open"] = True

    def edit_history_entry(self, event=None) -> None:
        tx_type = self.current_history_type
        selection = self.history_tree.selection()
        if not selection:
            messagebox.showinfo("안내", "수정할 내역을 선택해 주세요.")
            return
        if not selection[0].isdigit():
            return
        idx_in_view = int(selection[0])
        cache = self.history_cache.get(tx_type) or []
        index_map = self.history_indices.get(tx_type) or []
        if idx_in_view >= len(cache) or idx_in_view >= len(index_map):
            messagebox.showerror("오류", "선택한 기록을 찾을 수 없습니다.")
            return
        entry = cache[idx_in_view]
        dialog = TransactionEditDialog(self.root, entry, require_description=(tx_type == "out"))
        self.root.wait_window(dialog)
        if not dialog.result:
            return
        try:
            new_tx = Transaction(
                type=tx_type,
                artist=dialog.result["artist"],
                item=dialog.result["item"],
                category=self._normalize_category(dialog.result.get("category", "album")),
                option=dialog.result.get("option", ""),
                location=dialog.result["location"],
                quantity=int(dialog.result["quantity"]),
                timestamp=datetime.combine(date.fromisoformat(dialog.result["day"]), datetime.min.time()),
                actor=self._current_actor(),
                description=dialog.result.get("description", ""),
                event=entry.get("event", False),
                event_id=entry.get("event_id", ""),
                event_open=entry.get("event_open", False),
            )
            self._apply_history_edit(index_map[idx_in_view], new_tx)
        except ValueError as exc:
            messagebox.showerror("오류", str(exc))
            return
        self.set_status("내역을 수정했습니다.")
        self.refresh_stock()
        self._refresh_current_history()

    def delete_history_entry(self) -> None:
        tx_type = self.current_history_type
        selection = self.history_tree.selection()
        if not selection or not selection[0].isdigit():
            messagebox.showinfo("안내", "삭제할 내역을 선택해 주세요.")
            return
        idx_in_view = int(selection[0])
        cache = self.history_cache.get(tx_type) or []
        index_map = self.history_indices.get(tx_type) or []
        if idx_in_view >= len(cache) or idx_in_view >= len(index_map):
            messagebox.showerror("오류", "선택한 기록을 찾을 수 없습니다.")
            return
        entry = cache[idx_in_view]
        if not messagebox.askyesno("확인", "선택한 입/출고 기록을 삭제하시겠습니까?"):
            return
        history = self.data.get("history", [])
        entry_index = index_map[idx_in_view]
        if entry_index >= len(history):
            messagebox.showerror("오류", "선택한 기록을 찾을 수 없습니다.")
            return

        if entry.get("event") and entry.get("type") == "in" and entry.get("event_id"):
            self._reopen_event(entry.get("event_id"))

        qty = int(entry.get("quantity", 0))
        item = entry.get("item")
        option = entry.get("option", "")
        location = entry.get("location")

        if entry.get("type") == "in":
            available = (
                self.data.get("stock", {})
                .get(item, {})
                .get(option, {})
                .get(location, 0)
            )
            if available < qty:
                proceed = messagebox.askyesno(
                    "확인",
                    f"현재 재고({available}개)보다 적은 입고 삭제입니다. 음수 재고로 진행할까요?",
                )
                if not proceed:
                    return
            update_stock(self.data, item, option, location, -qty)
        else:
            update_stock(self.data, item, option, location, qty)

        history.pop(entry_index)
        self._save_async()
        self.refresh_stock()
        self._refresh_current_history()
        self.set_status("선택한 기록을 삭제했습니다.")
        self._log_user_action("입/출고 기록 삭제", persist=False)

    def clear_event_flag(self) -> None:
        selection = self.history_tree.selection()
        if not selection or not selection[0].isdigit():
            messagebox.showinfo("안내", "이벤트 출고 내역을 선택해 주세요.")
            return
        idx_in_view = int(selection[0])
        cache = self.history_cache.get(self.current_history_type) or []
        index_map = self.history_indices.get(self.current_history_type) or []
        if idx_in_view >= len(cache) or idx_in_view >= len(index_map):
            messagebox.showerror("오류", "선택한 기록을 찾을 수 없습니다.")
            return
        entry = cache[idx_in_view]
        if not (entry.get("event") and entry.get("type") == "out" and entry.get("event_open", False)):
            messagebox.showinfo("안내", "선택한 내역은 미해결 이벤트 출고가 아닙니다.")
            return
        history = self.data.get("history", [])
        hist_idx = index_map[idx_in_view]
        if hist_idx >= len(history):
            messagebox.showerror("오류", "선택한 기록을 찾을 수 없습니다.")
            return
        history[hist_idx]["event_open"] = False
        self._save_async()
        self.set_status("이벤트 색인을 해제했습니다.")
        self._refresh_current_history()

    def _apply_history_edit(self, entry_index: int, new_tx: Transaction) -> None:
        history = self.data.get("history", [])
        if entry_index >= len(history):
            raise ValueError("기록 인덱스가 올바르지 않습니다.")
        old = history[entry_index]

        if new_tx.type == "out" and not new_tx.description:
            raise ValueError("출고 시 상세내용은 필수입니다.")

        # 원본 영향 되돌리기
        if old.get("type") == "in":
            update_stock(
                self.data,
                old.get("item"),
                old.get("option", ""),
                old.get("location"),
                -int(old.get("quantity", 0)),
            )
        else:
            update_stock(
                self.data,
                old.get("item"),
                old.get("option", ""),
                old.get("location"),
                int(old.get("quantity", 0)),
            )

        available = (
            self.data.get("stock", {})
            .get(new_tx.item, {})
            .get(new_tx.option or "", {})
            .get(new_tx.location, 0)
        )
        if new_tx.type == "out" and available < new_tx.quantity:
            proceed = messagebox.askyesno(
                "확인",
                f"현재 재고({available}개)보다 많은 출고입니다. 음수 재고로 수정하시겠습니까?",
            )
            if not proceed:
                # 원복
                if old.get("type") == "in":
                    update_stock(
                        self.data,
                        old.get("item"),
                        old.get("option", ""),
                        old.get("location"),
                        int(old.get("quantity", 0)),
                    )
                else:
                    update_stock(
                        self.data,
                        old.get("item"),
                        old.get("option", ""),
                        old.get("location"),
                        -int(old.get("quantity", 0)),
                    )
                return

        ensure_period(self.data, new_tx.period)
        metadata = self.data.setdefault("item_metadata", {})
        meta_entry = metadata.setdefault(new_tx.item, {})
        meta_entry["artist"] = new_tx.artist
        meta_entry["category"] = new_tx.category
        if new_tx.option:
            meta_entry["option"] = new_tx.option

        if new_tx.type == "out":
            update_stock(self.data, new_tx.item, new_tx.option, new_tx.location, -new_tx.quantity)
        else:
            update_stock(self.data, new_tx.item, new_tx.option, new_tx.location, new_tx.quantity)

        history[entry_index] = new_tx.to_dict()
        self._save_async()
        self._log_user_action(
            f"입/출고 기록 수정 - {new_tx.type.upper()} {new_tx.item} {new_tx.option or '-'} @{new_tx.location} {new_tx.quantity}개",
            persist=False,
        )

    def export_history(self, tx_type: str) -> None:
        entries = self.history_cache.get(tx_type) or []
        if not entries:
            messagebox.showinfo("안내", "먼저 해당 유형으로 검색을 실행해 주세요.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All Files", "*.*")],
            title="엑셀 파일로 저장",
        )
        if not path:
            return
        try:
            directory = os.path.dirname(path)
            if directory:
                os.makedirs(directory, exist_ok=True)
            export_to_xlsx(entries, path, include_summary=True)
        except RuntimeError as exc:
            messagebox.showerror("오류", str(exc))
            return
        messagebox.showinfo("완료", f"엑셀 파일을 저장했습니다: {path}")

    # ---------------------------------------------------------------- google sync
    @staticmethod
    def _extract_sheet_id(raw: str) -> str:
        value = raw.strip()
        if not value:
            return ""
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", value)
        if match:
            return match.group(1)
        match = re.search(r"[?&]id=([a-zA-Z0-9-_]+)", value)
        if match:
            return match.group(1)
        return value

    def _get_google_sheet(self):
        if not self.settings.get("google_enabled", False):
            messagebox.showinfo("안내", "설정에서 구글 드라이브 연동을 활성화해 주세요.")
            return None
        sheet_raw = str(self.settings.get("google_sheet_id", "")).strip()
        sheet_id = self._extract_sheet_id(sheet_raw)
        creds_path = str(self.settings.get("google_credentials_path", "")).strip()
        if not sheet_id or not creds_path:
            messagebox.showinfo("안내", "설정에서 구글 시트 ID와 서비스 계정 JSON 경로를 입력해 주세요.")
            return None
        if not os.path.exists(creds_path):
            messagebox.showerror("오류", f"서비스 계정 JSON 파일을 찾을 수 없습니다: {creds_path}")
            return None
        try:
            import gspread
            from google.oauth2.service_account import Credentials
        except ImportError:
            messagebox.showerror(
                "오류",
                "구글 연동을 위해 gspread 및 google-auth 패키지가 필요합니다.\n"
                "pip install gspread google-auth 를 실행해 주세요.",
            )
            return None
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        try:
            creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
            client = gspread.authorize(creds)
            return client.open_by_key(sheet_id)
        except Exception as exc:  # pragma: no cover - external service
            status = getattr(getattr(exc, "response", None), "status_code", None)
            if status in (403, 404):
                messagebox.showerror(
                    "오류",
                    f"구글 시트 접근 권한이 없습니다. (HTTP {status})\n"
                    "서비스 계정이 시트에 공유되었는지 확인해 주세요.",
                )
            else:
                messagebox.showerror("오류", f"구글 시트 연결 실패: {exc}")
            return None

    @staticmethod
    def _get_or_create_sheet(worksheet_parent, title: str, rows: int = 1, cols: int = 1):
        sheets = worksheet_parent.worksheets()
        target = title.casefold()
        for ws in sheets:
            if ws.title.casefold() == target:
                print(f"[google-sync] Reusing worksheet '{ws.title}' for '{title}'")
                return ws
        print(f"[google-sync] Creating worksheet '{title}' (rows={rows}, cols={cols})")
        return worksheet_parent.add_worksheet(title=title, rows=rows, cols=cols)

    @staticmethod
    def _get_primary_sheet(worksheet_parent):
        try:
            return worksheet_parent.sheet1
        except Exception:
            sheets = worksheet_parent.worksheets()
            if sheets:
                return sheets[0]
        return worksheet_parent.add_worksheet(title="Stock", rows=1, cols=1)

    def _read_google_payload(self, sheet):
        sheets = {ws.title.casefold(): ws for ws in sheet.worksheets()}

        def get_ws(name: str, rows: int = 1, cols: int = 1):
            target = sheets.get(name.casefold())
            if target:
                return target
            return self._get_or_create_sheet(sheet, name, rows=rows, cols=cols)

        stock_ws_album = sheets.get("stock_album") or sheets.get("stock (album)")
        stock_ws_md = sheets.get("stock_md") or sheets.get("stock (md)")
        stock_ws_fallback = sheets.get("stock")
        history_ws = get_ws("History", rows=1, cols=1)
        meta_ws = get_ws("Metadata", rows=2, cols=2)

        meta_value = ""
        meta = meta_ws.get("A1:B1")
        if meta and meta[0] and len(meta[0]) >= 2 and meta[0][0] == "last_updated":
            meta_value = meta[0][1]

        stock_sources = []
        if stock_ws_album:
            stock_sources.append((stock_ws_album, "album"))
        if stock_ws_md:
            stock_sources.append((stock_ws_md, "md"))
        if not stock_sources and stock_ws_fallback:
            stock_sources.append((stock_ws_fallback, None))

        history_rows = history_ws.get_all_values() if history_ws else []
        if not stock_sources and not history_rows:
            return None, meta_value

        data = {
            "current_period": None,
            "periods": {},
            "stock": {},
            "history": [],
            "item_metadata": {},
            "last_updated": meta_value or None,
        }

        def normalize_stock_header(value: str) -> str:
            lowered = value.strip().lower()
            if "아티스트" in value or "artist" in lowered:
                return "artist"
            if "앨범" in value or "item" in lowered or "album" in lowered:
                return "item"
            if "옵션" in value or "option" in lowered:
                return "option"
            if "현재고" in value or "수량" in value or "quantity" in lowered:
                return "quantity"
            if "로케이션" in value or "location" in lowered:
                return "location"
            if "구분" in value or "category" in lowered:
                return "category"
            return lowered

        for ws, category_hint in stock_sources:
            rows = ws.get_all_values()
            if not rows:
                continue
            headers = [normalize_stock_header(h) for h in rows[0]]
            for row in rows[1:]:
                if not any(cell.strip() for cell in row):
                    continue
                values = dict(zip(headers, row))
                item = values.get("item") or ""
                artist = values.get("artist") or ""
                option = values.get("option") or ""
                location = values.get("location") or ""
                qty_raw = values.get("quantity") or "0"
                category_value = normalize_category(values.get("category") or category_hint or "album")
                try:
                    qty = int(str(qty_raw).replace(",", ""))
                except ValueError:
                    qty = 0
                if not item:
                    continue
                option_key = option or ""
                data["stock"].setdefault(item, {}).setdefault(option_key, {})[location] = qty
                meta_entry = data["item_metadata"].setdefault(item, {})
                if artist:
                    meta_entry["artist"] = artist
                meta_entry.setdefault("category", category_value)

        if history_rows:
            headers = [h.strip().lower() for h in history_rows[0]]
            for row in history_rows[1:]:
                if not any(cell.strip() for cell in row):
                    continue
                values = dict(zip(headers, row))
                tx_type = values.get("type") or values.get("타입") or ""
                artist = values.get("artist") or values.get("아티스트") or ""
                item = values.get("item") or values.get("앨범/버전") or ""
                option = values.get("option") or values.get("옵션") or ""
                location = values.get("location") or values.get("로케이션") or ""
                qty_raw = values.get("quantity") or values.get("수량") or "0"
                timestamp = values.get("timestamp") or values.get("기록시각") or ""
                day = values.get("day") or values.get("일자") or ""
                period = values.get("period") or values.get("월") or ""
                year = values.get("year") or values.get("연") or ""
                description = values.get("description") or values.get("상세내용") or ""
                actor = values.get("actor") or values.get("작성자") or ""
                category_value = normalize_category(values.get("category") or values.get("구분") or "album")
                event_raw = values.get("event") or values.get("이벤트") or ""
                event_id = values.get("event_id") or values.get("이벤트id") or ""
                event_open_raw = values.get("event_open") or values.get("이벤트열림") or ""
                try:
                    qty = int(str(qty_raw).replace(",", ""))
                except ValueError:
                    qty = 0
                if not item or not tx_type:
                    continue
                data["history"].append(
                    {
                        "type": tx_type,
                        "artist": artist,
                        "item": item,
                        "category": category_value,
                        "option": option,
                        "location": location,
                        "quantity": qty,
                        "timestamp": timestamp,
                        "day": day,
                        "period": period,
                        "year": year,
                        "description": description,
                        "actor": actor,
                        "event": str(event_raw).lower() in {"true", "1", "y", "yes"},
                        "event_id": event_id,
                        "event_open": str(event_open_raw).lower() in {"true", "1", "y", "yes"},
                    }
                )
                meta_entry = data["item_metadata"].setdefault(item, {})
                if artist:
                    meta_entry["artist"] = artist
                meta_entry.setdefault("category", category_value)

        if data["history"]:
            periods = sorted({entry.get("period") for entry in data["history"] if entry.get("period")})
            data["current_period"] = periods[-1] if periods else None
        return data, meta_value

    def _write_google_payload(self, sheet, data: Dict) -> None:
        stock_album_ws = self._get_or_create_sheet(sheet, "Stock_Album", rows=1, cols=1)
        stock_md_ws = self._get_or_create_sheet(sheet, "Stock_MD", rows=1, cols=1)
        history_ws = self._get_or_create_sheet(sheet, "History", rows=1, cols=1)
        meta_ws = self._get_or_create_sheet(sheet, "Metadata", rows=2, cols=2)

        album_rows = [["아티스트", "앨범/버전", "옵션", "현재고", "로케이션"]]
        md_rows = [["아티스트", "앨범/버전", "옵션", "현재고", "로케이션"]]
        for item, options in sorted((data.get("stock") or {}).items()):
            meta_entry = data.get("item_metadata", {}).get(item, {})
            artist = meta_entry.get("artist", "")
            category_value = normalize_category(meta_entry.get("category", "album"))
            target_rows = md_rows if category_value == "md" else album_rows
            for option, locations in sorted(options.items()):
                for location, qty in sorted(locations.items()):
                    target_rows.append([artist, item, option or "", str(qty), location])

        history_values = [
            [
                "Type",
                "Artist",
                "Category",
                "Item",
                "Option",
                "Location",
                "Quantity",
                "Timestamp",
                "Day",
                "Period",
                "Year",
                "Description",
                "Actor",
                "Event",
                "EventId",
                "EventOpen",
            ]
        ]
        for entry in data.get("history", []):
            history_values.append(
                [
                    entry.get("type", ""),
                    entry.get("artist", ""),
                    normalize_category(entry.get("category", "album")),
                    entry.get("item", ""),
                    entry.get("option", ""),
                    entry.get("location", ""),
                    str(entry.get("quantity", "")),
                    self._format_history_timestamp(entry),
                    entry.get("day", ""),
                    entry.get("period", ""),
                    entry.get("year", ""),
                    entry.get("description", ""),
                    entry.get("actor", ""),
                    str(entry.get("event", False)),
                    entry.get("event_id", ""),
                    str(entry.get("event_open", False)),
                ]
            )

        meta_values = [["last_updated", data.get("last_updated") or ""]]
        try:
            stock_album_ws.update("A1", album_rows if album_rows else [[""]])
            stock_md_ws.update("A1", md_rows if md_rows else [[""]])
            history_ws.update("A1", history_values if history_values else [[""]])
            meta_ws.update("A1:B1", meta_values)
        except Exception as exc:  # pragma: no cover - external service
            sheet_id = getattr(sheet, "id", None)
            sheet_name = getattr(history_ws, "title", "unknown")
            raise RuntimeError(
                "구글 시트 저장 실패: "
                f"SpreadsheetId={sheet_id}, Sheet={sheet_name}, values_shape={len(history_values)} rows"
            ) from exc

    @staticmethod
    def _parse_timestamp(raw: Optional[str]) -> Optional[datetime]:
        if not raw:
            return None
        try:
            return datetime.fromisoformat(str(raw))
        except Exception:
            return None

    @staticmethod
    def _format_history_timestamp(entry: Dict[str, object]) -> str:
        raw = str(entry.get("timestamp") or "").strip()
        if raw:
            normalized = raw.replace("Z", "")
            parsed = InventoryApp._parse_timestamp(normalized)
            if parsed:
                return parsed.strftime("%Y-%m-%d %H:%M:%S")
            if len(normalized) == 10 and normalized.count("-") == 2:
                return f"{normalized} 00:00:00"
            return normalized
        day = str(entry.get("day") or "").strip()
        if day:
            return f"{day} 00:00:00"
        return ""

    def _confirm_action(self, title: str, message: str, yes_label: str, no_label: str) -> bool:
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("520x220")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        result = tk.BooleanVar(value=False)

        frame = ttk.Frame(dialog, padding=16)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=message, wraplength=460, justify=tk.LEFT).pack(anchor="w")

        btns = ttk.Frame(frame)
        btns.pack(anchor="e", pady=(20, 0))

        def choose(value: bool) -> None:
            result.set(value)
            dialog.destroy()

        ttk.Button(btns, text=no_label, command=lambda: choose(False)).pack(side=tk.RIGHT)
        ttk.Button(btns, text=yes_label, command=lambda: choose(True)).pack(side=tk.RIGHT, padx=(0, 8))

        dialog.wait_window()
        return result.get()

    @staticmethod
    def _build_stock_snapshot(data: Dict) -> Dict[Tuple[str, str, str], Dict[str, object]]:
        snapshot: Dict[Tuple[str, str, str, str], Dict[str, object]] = {}
        metadata = data.get("item_metadata", {})
        for item, options in (data.get("stock") or {}).items():
            artist = metadata.get(item, {}).get("artist", "")
            category = normalize_category(metadata.get(item, {}).get("category", "album"))
            for option, locations in options.items():
                for location, qty in locations.items():
                    key = (category, item, option or "", location or "")
                    snapshot[key] = {"qty": int(qty), "artist": artist, "category": category}
        return snapshot

    def sync_google_drive(self) -> None:
        """Main toolbar sync: always pull Google Sheets into the app."""

        confirmed = self._confirm_action(
            "구글 드라이브 → Inventory Manager",
            "주의 : 현재 구글 시트의 데이터를 inventory manager로 업데이트 진행하는것을 인지하고 계신가요?\n"
            "(업데이트 완료 후에는 돌이킬 수 없습니다.)",
            "인지하고 있습니다",
            "아니요",
        )
        if not confirmed:
            return
        sheet = self._get_google_sheet()
        if sheet is None:
            return
        try:
            google_data, _google_updated_raw = self._read_google_payload(sheet)
        except Exception as exc:
            messagebox.showerror("오류", str(exc))
            return
        if not google_data:
            messagebox.showinfo("안내", "구글 시트에 데이터가 없습니다. 먼저 초기 업로드를 진행해 주세요.")
            return
        backup_data_with_label(self.data, "google_pull")

        local_snapshot = self._build_stock_snapshot(self.data)
        google_snapshot = self._build_stock_snapshot(google_data)
        changes = 0

        for item, meta in google_data.get("item_metadata", {}).items():
            if meta.get("artist"):
                self.data.setdefault("item_metadata", {}).setdefault(item, {})["artist"] = meta["artist"]

        all_keys = sorted(set(local_snapshot) | set(google_snapshot))
        now = datetime.now()
        for key in all_keys:
            local_info = local_snapshot.get(key, {"qty": 0, "artist": "", "category": "album"})
            google_info = google_snapshot.get(key, {"qty": 0, "artist": "", "category": "album"})
            local_qty = int(local_info.get("qty", 0))
            google_qty = int(google_info.get("qty", 0))
            diff = google_qty - local_qty
            if diff == 0:
                continue
            category, item, option, location = key
            artist = (
                google_info.get("artist")
                or local_info.get("artist")
                or determine_artist(self.data, item)
                or ""
            )
            category_value = normalize_category(google_info.get("category") or local_info.get("category") or "album")
            tx_type = "in" if diff > 0 else "out"
            transaction = Transaction(
                type=tx_type,
                artist=artist,
                item=item,
                category=category_value,
                option=option,
                location=location,
                quantity=abs(diff),
                timestamp=now,
                actor=self._current_actor(),
                description="Google Drive 수정",
            )
            try:
                record_transaction(self.data, transaction, allow_negative=True)
            except Exception as exc:
                messagebox.showerror("오류", f"구글 시트 반영 중 오류가 발생했습니다: {exc}")
                return
            changes += 1

        if changes == 0:
            messagebox.showinfo("안내", "구글 시트와 로컬 데이터가 동일합니다.")
            return

        self._save_now()
        self.reload_data()
        self.set_status("구글 드라이브 데이터를 반영했습니다.")
        self._log_user_action("구글 시트 → Inventory Manager 동기화", persist=True)

    def upload_google_drive(self) -> None:
        """Settings-only: always push local data to Google Sheets."""

        confirmed = self._confirm_action(
            "Inventory Manager → 구글 드라이브",
            "주의 : 현재 inventory manager의 데이터를 구글 시트로 업데이트 진행하는것을 인지하고 계신가요?\n"
            "(업데이트 완료 후에는 돌이킬 수 없습니다.)",
            "인지하고 있습니다",
            "아니요",
        )
        if not confirmed:
            return
        sheet = self._get_google_sheet()
        if sheet is None:
            return
        backup_data_with_label(self.data, "google_push")
        self._save_now()
        try:
            self._write_google_payload(sheet, self.data)
        except Exception as exc:
            messagebox.showerror("오류", f"구글 시트 저장 실패: {exc}")
            return
        self.set_status("로컬 데이터를 구글 드라이브로 업로드했습니다.")
        self._log_user_action("Inventory Manager → 구글 시트 동기화", persist=True)

    # ------------------------------------------------------------------ 잠금 및 설정
    def _start_idle_watch(self) -> None:
        """Begin monitoring idle time based on configured timeout."""

        if self._idle_job:
            self.root.after_cancel(self._idle_job)
        self.last_activity = datetime.now()
        self._idle_job = self.root.after(5000, self._check_idle)

    def _reset_idle_timer(self) -> None:
        self.last_activity = datetime.now()
        if self._idle_job:
            self.root.after_cancel(self._idle_job)
        self._idle_job = self.root.after(5000, self._check_idle)

    def _check_idle(self) -> None:
        timeout_min = int(self.settings.get("idle_minutes", 0) or 0)
        should_lock = (
            self.settings.get("lock_on_idle", True)
            and timeout_min > 0
            and (datetime.now() - self.last_activity) >= timedelta(minutes=timeout_min)
        )
        if should_lock:
            self._show_lock_dialog("사용자 활동이 없어 잠금되었습니다.")
        self._idle_job = self.root.after(5000, self._check_idle)

    def _update_activity(self, *_args) -> None:
        if self._lock_window is not None:
            return
        self.last_activity = datetime.now()

    def _maybe_lock_on_start(self) -> None:
        if self.settings.get("lock_on_start", True):
            self.root.after(100, lambda: self._show_lock_dialog("시작 시 잠금이 적용되었습니다."))

    def _show_lock_dialog(self, reason: str) -> None:
        if self._lock_window is not None:
            return

        self._lock_window = tk.Toplevel(self.root)
        self._lock_window.title("잠금 모드")
        self._lock_window.geometry("420x220")
        self._lock_window.transient(self.root)
        self._lock_window.grab_set()
        self._lock_window.protocol("WM_DELETE_WINDOW", lambda: None)

        frame = ttk.Frame(self._lock_window, padding=16)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=reason, foreground="#b45309").pack(anchor="w")
        ttk.Label(frame, text="비밀번호를 입력하세요.", padding=(0, 6)).pack(anchor="w")
        pwd_var = tk.StringVar()
        entry = ttk.Entry(frame, textvariable=pwd_var, show="*")
        entry.pack(fill=tk.X)
        entry.focus_set()

        def unlock() -> None:
            expected = str(self.settings.get("password", ""))
            if expected and pwd_var.get() != expected:
                messagebox.showerror("오류", "비밀번호가 올바르지 않습니다.", parent=self._lock_window)
                return
            if self._lock_window:
                self._lock_window.grab_release()
                self._lock_window.destroy()
            self._lock_window = None
            self._reset_idle_timer()

        ttk.Button(frame, text="잠금 해제", command=unlock).pack(pady=(12, 0))

    def open_settings_window(self) -> None:
        win = tk.Toplevel(self.root)
        win.title("설정")
        win.geometry("560x520")
        win.transient(self.root)
        win.grab_set()

        pwd_var = tk.StringVar(value=str(self.settings.get("password", "")))
        confirm_var = tk.StringVar(value=str(self.settings.get("password", "")))
        idle_var = tk.StringVar(value=str(self.settings.get("idle_minutes", 5)))
        lock_start_var = tk.BooleanVar(value=bool(self.settings.get("lock_on_start", True)))
        lock_idle_var = tk.BooleanVar(value=bool(self.settings.get("lock_on_idle", True)))
        nickname_var = tk.StringVar(value=str(self.settings.get("nickname", "")))
        nickname_pwd_var = tk.StringVar()
        nickname_pwd_confirm_var = tk.StringVar()
        nickname_pwd_current_var = tk.StringVar()
        google_enabled_var = tk.BooleanVar(value=bool(self.settings.get("google_enabled", False)))
        google_sheet_var = tk.StringVar(value=str(self.settings.get("google_sheet_id", "")))
        google_creds_var = tk.StringVar(value=str(self.settings.get("google_credentials_path", "")))
        location_presets: List[str] = list(self.settings.get("location_presets", []) or [])

        body = ttk.Frame(win, padding=12)
        body.pack(fill=tk.BOTH, expand=True)

        ttk.Label(body, text="닉네임").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(body, textvariable=nickname_var).grid(row=0, column=1, sticky=tk.EW, padx=(8, 0))
        ttk.Label(body, text="닉네임 변경 암호 (새로 설정/변경 시)").grid(row=1, column=0, sticky=tk.W, pady=(6, 0))
        ttk.Entry(body, textvariable=nickname_pwd_var, show="*").grid(row=1, column=1, sticky=tk.EW, padx=(8, 0), pady=(6, 0))
        ttk.Label(body, text="닉네임 암호 확인").grid(row=2, column=0, sticky=tk.W)
        ttk.Entry(body, textvariable=nickname_pwd_confirm_var, show="*").grid(row=2, column=1, sticky=tk.EW, padx=(8, 0))
        ttk.Label(body, text="현재 닉네임 암호 (변경 시 인증)").grid(row=3, column=0, sticky=tk.W, pady=(6, 0))
        ttk.Entry(body, textvariable=nickname_pwd_current_var, show="*").grid(row=3, column=1, sticky=tk.EW, padx=(8, 0), pady=(6, 0))

        ttk.Label(body, text="새 비밀번호").grid(row=4, column=0, sticky=tk.W, pady=(6, 0))
        ttk.Entry(body, textvariable=pwd_var, show="*").grid(row=4, column=1, sticky=tk.EW, padx=(8, 0))
        ttk.Label(body, text="비밀번호 확인").grid(row=5, column=0, sticky=tk.W, pady=(6, 0))
        ttk.Entry(body, textvariable=confirm_var, show="*").grid(row=5, column=1, sticky=tk.EW, padx=(8, 0), pady=(6, 0))
        ttk.Label(body, text="자동 잠금 대기(분)").grid(row=6, column=0, sticky=tk.W, pady=(6, 0))
        ttk.Entry(body, textvariable=idle_var).grid(row=6, column=1, sticky=tk.EW, padx=(8, 0), pady=(6, 0))
        ttk.Checkbutton(body, text="프로그램 실행 시 잠금", variable=lock_start_var).grid(
            row=7, column=0, columnspan=2, sticky=tk.W, pady=(6, 0)
        )
        ttk.Checkbutton(body, text="입력 없음 시 자동 잠금", variable=lock_idle_var).grid(
            row=8, column=0, columnspan=2, sticky=tk.W, pady=(4, 0)
        )

        body.columnconfigure(1, weight=1)

        ttk.Separator(win, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(8, 0))

        google_frame = ttk.Frame(win, padding=12)
        google_frame.pack(fill=tk.X)
        ttk.Label(google_frame, text="구글 드라이브 연동").grid(row=0, column=0, sticky=tk.W)
        ttk.Checkbutton(google_frame, text="연동 사용", variable=google_enabled_var).grid(
            row=0, column=1, sticky=tk.W, padx=(8, 0)
        )
        ttk.Label(google_frame, text="구글 시트 ID").grid(row=1, column=0, sticky=tk.W, pady=(6, 0))
        ttk.Entry(google_frame, textvariable=google_sheet_var).grid(
            row=1, column=1, sticky=tk.EW, padx=(8, 0), pady=(6, 0)
        )
        ttk.Label(google_frame, text="서비스 계정 JSON").grid(row=2, column=0, sticky=tk.W, pady=(6, 0))
        creds_entry = ttk.Entry(google_frame, textvariable=google_creds_var)
        creds_entry.grid(row=2, column=1, sticky=tk.EW, padx=(8, 0), pady=(6, 0))

        def browse_google_creds() -> None:
            path = filedialog.askopenfilename(
                title="구글 서비스 계정 JSON 선택",
                filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")],
            )
            if path:
                google_creds_var.set(path)

        ttk.Button(google_frame, text="찾아보기", command=browse_google_creds).grid(
            row=2, column=2, padx=(8, 0), pady=(6, 0)
        )
        google_frame.columnconfigure(1, weight=1)

        ttk.Separator(win, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(4, 0))

        preset_frame = ttk.Frame(win, padding=12)
        preset_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(preset_frame, text="로케이션 프리셋 (드롭다운/선택 목록)").grid(row=0, column=0, sticky=tk.W)
        preset_frame.columnconfigure(0, weight=1)

        preset_controls = ttk.Frame(preset_frame)
        preset_controls.grid(row=1, column=0, sticky=tk.EW, pady=(6, 6))
        preset_controls.columnconfigure(0, weight=1)
        preset_var = tk.StringVar()
        preset_entry = ttk.Entry(preset_controls, textvariable=preset_var)
        preset_entry.grid(row=0, column=0, sticky=tk.EW)

        preset_buttons = ttk.Frame(preset_controls)
        preset_buttons.grid(row=0, column=1, padx=(8, 0))
        ttk.Button(preset_buttons, text="추가", command=lambda: add_preset()).pack(side=tk.LEFT)
        ttk.Button(preset_buttons, text="삭제", command=lambda: remove_selected()).pack(side=tk.LEFT, padx=(4, 0))

        preset_list = tk.Listbox(preset_frame, height=6)
        preset_list.grid(row=2, column=0, sticky=tk.NSEW)
        preset_frame.rowconfigure(2, weight=1)

        for value in location_presets:
            preset_list.insert(tk.END, value)

        def add_preset() -> None:
            value = preset_var.get().strip()
            if not value:
                return
            existing = [preset_list.get(idx) for idx in range(preset_list.size())]
            if value in existing:
                preset_var.set("")
                return
            preset_list.insert(tk.END, value)
            preset_var.set("")

        def remove_selected() -> None:
            selection = preset_list.curselection()
            if not selection:
                return
            for idx in reversed(selection):
                preset_list.delete(idx)

        def save_and_close() -> None:
            if pwd_var.get() != confirm_var.get():
                messagebox.showerror("오류", "비밀번호가 일치하지 않습니다.", parent=win)
                return
            if nickname_pwd_var.get() and nickname_pwd_var.get() != nickname_pwd_confirm_var.get():
                messagebox.showerror("오류", "닉네임 암호가 일치하지 않습니다.", parent=win)
                return
            try:
                minutes = int(idle_var.get())
                if minutes < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("오류", "자동 잠금 시간은 0 이상의 숫자여야 합니다.", parent=win)
                return

            current_nick_pwd = str(self.settings.get("nickname_password", ""))
            new_nick = nickname_var.get().strip()
            new_nick_pwd = nickname_pwd_var.get() or current_nick_pwd
            if (new_nick != self.settings.get("nickname", "") or new_nick_pwd != current_nick_pwd) and current_nick_pwd:
                if nickname_pwd_current_var.get() != current_nick_pwd:
                    messagebox.showerror("오류", "현재 닉네임 암호가 올바르지 않습니다.", parent=win)
                    return
            if new_nick_pwd and pwd_var.get() and new_nick_pwd == pwd_var.get():
                messagebox.showerror("오류", "닉네임 암호와 잠금 비밀번호는 달라야 합니다.", parent=win)
                return

            self.settings = {
                "password": pwd_var.get(),
                "lock_on_start": lock_start_var.get(),
                "lock_on_idle": lock_idle_var.get(),
                "idle_minutes": minutes,
                "location_presets": [preset_list.get(i) for i in range(preset_list.size())],
                "google_enabled": google_enabled_var.get(),
                "google_sheet_id": google_sheet_var.get().strip(),
                "google_credentials_path": google_creds_var.get().strip(),
                "nickname": new_nick,
                "nickname_password": new_nick_pwd,
            }
            save_settings(self.settings)
            self._apply_location_presets()
            self._reset_idle_timer()
            messagebox.showinfo("완료", "설정을 저장했습니다.", parent=win)
            win.destroy()

        btns = ttk.Frame(win, padding=(12, 0, 12, 12))
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="잠금", command=lambda: self._show_lock_dialog("수동으로 잠금이 실행되었습니다.")).pack(
            side=tk.LEFT
        )
        ttk.Button(btns, text="취소", command=win.destroy).pack(side=tk.RIGHT)
        ttk.Button(btns, text="저장", command=save_and_close).pack(side=tk.RIGHT, padx=(0, 8))

    def show_artist_summary(self) -> None:
        window = tk.Toplevel(self.root)
        window.title("아티스트별 재고 상세")
        window.geometry("820x460")
        frame = ttk.Frame(window, padding=8)
        frame.pack(fill=tk.BOTH, expand=True)

        columns = ("artist", "item", "option", "opening", "in_total", "out_total", "qty", "audit", "location")
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        headings = [
            ("artist", "아티스트", 140),
            ("item", "앨범/버전", 200),
            ("option", "옵션", 120),
            ("opening", "기초재고", 90),
            ("in_total", "입고합계", 90),
            ("out_total", "출고합계", 90),
            ("qty", "현재고", 80),
            ("audit", "마지막 실사", 110),
            ("location", "로케이션", 140),
        ]
        for col, title, width in headings:
            tree.heading(col, text=title)
            tree.column(col, width=width, anchor=tk.E if col in {"opening", "in_total", "out_total", "qty"} else tk.W)
        for row in sorted(self.stock_rows, key=lambda r: (r[0], r[1], r[8])):
            tree.insert("", tk.END, values=row)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=scrollbar.set)


def main() -> None:
    try:
        root = tk.Tk()
        apply_modern_styles(root)
    except Exception as exc:  # pragma: no cover - startup safety
        _log_fatal("Tk init failed", exc)
        print(f"Tk 초기화 실패: {exc}")
        return

    try:
        app = InventoryApp(root)
    except Exception as exc:  # pragma: no cover - startup safety
        _log_fatal("App startup failed", exc)
        messagebox.showerror(
            "시작 오류",
            "프로그램을 시작할 수 없습니다.\n"
            f"{exc}\n"
            f"자세한 내용은 {FATAL_LOG} 파일을 확인하세요.",
        )
        root.destroy()
        return

    root.mainloop()


if __name__ == "__main__":
    main()
